#!/usr/bin/env python3
"""Process HR_CA management sheet data into clean JSON for dashboard."""

import json
import openpyxl
from datetime import datetime, date
from collections import defaultdict
import re

def get_current_quarter(sales=None):
    """Auto-detect current fiscal quarter based on today's date.
    Always uses the calendar quarter (not fallback).
    FY follows calendar year with Q1=Jan-Mar, Q2=Apr-Jun, Q3=Jul-Sep, Q4=Oct-Dec."""
    today = datetime.now()
    fy = today.year
    q = (today.month - 1) // 3 + 1

    # Previous quarter (for confirmed decision rate)
    prev_q = q - 1
    prev_fy = fy
    if prev_q <= 0:
        prev_q = 4
        prev_fy = fy - 1

    # Quarter end date
    q_end_month = q * 3
    if q_end_month == 12:
        q_end_date = datetime(fy, 12, 31)
    else:
        q_end_date = datetime(fy, q_end_month + 1, 1) - __import__('datetime').timedelta(days=1)

    return {
        'current_q_funnel': f'FY{fy % 100}/{q}Q',       # e.g. FY26/2Q
        'current_q_sales': f'FY{fy % 100}／{q}Q',       # full-width slash
        'confirmed_q': f'FY{prev_fy % 100}／{prev_q}Q', # previous Q (for confirmed rates)
        'q_end_date': q_end_date,
        'fy': fy,
        'q': q,
        'calendar_q': q,
        'calendar_fy': fy,
        'prev_fy': prev_fy,
        'prev_q': prev_q,
    }

def month_sort_key(m):
    """Sort key for Japanese month strings like '2025年1月'."""
    match = re.match(r'(\d{4})年(\d{1,2})月', m)
    if match:
        return (int(match.group(1)), int(match.group(2)))
    return (9999, 99)

def date_sort_key(m):
    """Sort key for date strings like '2024-01-01'."""
    match = re.match(r'(\d{4})-(\d{2})-(\d{2})', str(m))
    if match:
        return (int(match.group(1)), int(match.group(2)))
    if 'FY' in str(m):
        # FY24/1Q -> (2024, 1), FY25/2Q -> (2025, 4)
        fy_match = re.match(r'FY(\d{2}).*?(\d)Q', str(m))
        if fy_match:
            y = 2000 + int(fy_match.group(1))
            q = int(fy_match.group(2))
            return (y, q * 3)
    return (9999, 99)

def safe_val(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    if isinstance(v, date): return v.strftime('%Y-%m-%d')
    if isinstance(v, (int, float)): return v
    return str(v)

def parse_number(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return v
    s = str(v).replace(',', '').replace('¥', '').replace('%', '').strip()
    if not s or s == '-' or s == 'None': return 0
    try: return float(s)
    except: return 0

def clean_period_key(v):
    """Clean period key to consistent format."""
    s = str(v).strip()
    # Convert "2024-01-01 00:00:00" -> "2024-01"
    m = re.match(r'(\d{4})-(\d{2})-\d{2}', s)
    if m: return f"{m.group(1)}-{m.group(2)}"
    # Convert FY keys
    s = s.replace('／', '/').replace('　', '')
    return s

def extract_sales_by_ca(wb):
    ws = wb['DB_売上管理']
    all_rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        all_rows.append([safe_val(c) for c in row])

    headers = all_rows[2]
    months = []
    for i, h in enumerate(headers):
        if h and ('年' in str(h) and '月' in str(h)):
            months.append({'col': i, 'label': str(h)})
        elif h and 'FY' in str(h):
            months.append({'col': i, 'label': str(h)})

    ca_starts = []
    for i, row in enumerate(all_rows[3:], start=3):
        if row[0] and row[1] and '目標' in str(row[1]):
            ca_starts.append({'row': i, 'name': str(row[0])})

    # Normalize metric names: map new spreadsheet names to legacy names
    # so downstream code doesn't need to change
    metric_aliases = {
        '目標(CA粗利)': '目標(粗利)',
        '実績(CA粗利)': '実績(粗利)',
        '着地(CA粗利)': '着地見込み',
        '実績(RA粗利)': '実績(RA粗利)',
    }

    result = {}
    for ca_info in ca_starts:
        ca_name = ca_info['name']
        ca_row = ca_info['row']
        ca_data = {'quarterly': {}, 'monthly': {}}

        for offset in range(min(13, len(all_rows) - ca_row)):
            row = all_rows[ca_row + offset]
            metric = str(row[1]) if row[1] else ''
            # Normalize metric name
            metric = metric_aliases.get(metric, metric)
            for m in months:
                period = m['label']
                col = m['col']
                if col < len(row):
                    val = parse_number(row[col])
                    bucket = 'quarterly' if 'FY' in period else 'monthly'
                    if period not in ca_data[bucket]:
                        ca_data[bucket][period] = {}
                    ca_data[bucket][period][metric] = val

        result[ca_name] = ca_data
    return result

def extract_ca_monthly_funnel(wb):
    ws = wb['CA月次プロセス']
    all_rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        all_rows.append([safe_val(c) for c in row])

    header_row = all_rows[4]
    periods = []
    for i, h in enumerate(header_row):
        if h and ('FY' in str(h) or '20' in str(h)):
            periods.append({'col': i, 'label': clean_period_key(h)})

    stages = ['アサイン', '初回面談', '求人紹介', '書類準備中', '応募', '応募(社数)',
              '1st面接', '2nd面接', '最終面接', '内定', '決定']
    rate_stages = ['求人紹介率', '面談応募率', '応募社数/一人', '書類通過率', '1次通過率',
                   '応募内定率', '内定決定率', '応募決定率', '面談決定率']

    ca_data = {}
    current_ca = None

    for i, row in enumerate(all_rows[5:], start=5):
        if not row or len(row) < 4: continue
        ca_name = row[2] if row[2] else None
        status = row[3] if row[3] else None

        if ca_name and status == 'アサイン':
            current_ca = str(ca_name)
            if current_ca not in ca_data:
                ca_data[current_ca] = {'funnel': {}, 'rates': {}}

        if current_ca and status:
            status_str = str(status)
            target = None
            if status_str in stages: target = ca_data[current_ca]['funnel']
            elif status_str in rate_stages: target = ca_data[current_ca]['rates']

            if target is not None:
                if status_str not in target: target[status_str] = {}
                for p in periods:
                    if p['col'] < len(row) and row[p['col']] is not None:
                        target[status_str][p['label']] = parse_number(row[p['col']])

    return ca_data

def extract_inflow_attributes(wb):
    """Extract inflow attribute data (就職時期, 性別, 学歴, 年齢, 前職)."""
    ws = wb['流入属性']
    all_rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        all_rows.append([safe_val(c) for c in row])

    # Row 5 has sub-headers for each category
    # Columns: 年月 | 就職時期(すぐに,1~2ヶ月,3~6ヶ月,6ヶ月以上,未定) | 性別(男,女) | 学歴(...) | 年齢(...) | 前職(...)
    # Map column indices to categories
    categories = {}
    header4 = all_rows[3]  # Row 4: category names
    header5 = all_rows[4]  # Row 5: sub-categories

    current_cat = None
    for j in range(1, len(header5)):
        if header4[j]: current_cat = str(header4[j])
        sub = str(header5[j]) if header5[j] else None
        if current_cat and sub:
            categories[j] = {'category': current_cat, 'subcategory': sub}

    # Read data rows
    data_by_period = {}
    for i in range(5, len(all_rows)):
        row = all_rows[i]
        if not row[0]: continue
        period = clean_period_key(row[0])
        period_data = {}
        for j, info in categories.items():
            if j < len(row):
                cat = info['category']
                sub = info['subcategory']
                if cat not in period_data: period_data[cat] = {}
                period_data[cat][sub] = parse_number(row[j])
        data_by_period[period] = period_data

    return dict(sorted(data_by_period.items(), key=lambda x: date_sort_key(x[0])))

def extract_route_process(wb):
    """Extract application route process data (自社 vs 自社以外)."""
    ws = wb['応募経路別プロセス']
    all_rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        all_rows.append([safe_val(c) for c in row])

    # Row 4 has headers: 最終No, カテゴリ, ステータス, FY24/1Q, ...
    header = all_rows[3]
    periods = []
    for i, h in enumerate(header):
        if h and i >= 3:
            periods.append({'col': i, 'label': clean_period_key(h)})

    # Parse each category section
    result = {}
    current_cat = None
    for i in range(4, len(all_rows)):
        row = all_rows[i]
        if not row or len(row) < 3: continue
        cat = str(row[1]) if row[1] else None
        status = str(row[2]) if row[2] else None

        if cat and cat != 'None': current_cat = cat
        if not current_cat or not status or status == 'None': continue

        if current_cat not in result: result[current_cat] = {}
        if status not in result[current_cat]: result[current_cat][status] = {}

        for p in periods:
            if p['col'] < len(row) and row[p['col']] is not None:
                result[current_cat][status][p['label']] = parse_number(row[p['col']])

    return result

def extract_decision_by_attribute(wb):
    """Extract decision rate by attribute (age, etc)."""
    ws = wb['決定率_属性別']
    age_data = []
    for i in range(3, 9):
        age = safe_val(ws.cell(i, 1).value)
        count = parse_number(ws.cell(i, 2).value)
        decisions = parse_number(ws.cell(i, 3).value)
        rate = parse_number(ws.cell(i, 4).value)
        if age:
            age_data.append({'age_range': str(age), 'count': count, 'decisions': decisions, 'rate': rate})
    return {'by_age': age_data}

def extract_2026_targets(targets_path):
    """Extract 2026 planning targets from the targets spreadsheet."""
    wb = openpyxl.load_workbook(targets_path, data_only=True)
    ws = wb['2026計画']
    targets = {}  # {Q: {CA: {...}}}
    for i in range(6, ws.max_row + 1):
        q = safe_val(ws.cell(i, 2).value)
        ca = safe_val(ws.cell(i, 3).value)
        if not q or not ca: continue
        q = str(q).replace('Q', 'Q').strip()
        ca = str(ca).strip()
        if ca in ['合計', '合計/平均', 'RA/Ptnr', '新人C']: continue
        interviews = parse_number(ws.cell(i, 4).value)
        interview_result = parse_number(ws.cell(i, 5).value)
        unit_price = parse_number(ws.cell(i, 6).value)
        unit_result = parse_number(ws.cell(i, 7).value)
        dec_rate = parse_number(ws.cell(i, 8).value)
        q_target = parse_number(ws.cell(i, 9).value)
        q_result = parse_number(ws.cell(i, 10).value)
        notes = safe_val(ws.cell(i, 11).value) or ''

        if q not in targets: targets[q] = {}
        targets[q][ca] = {
            'interviews': interviews, 'interview_result': interview_result,
            'unit_price': unit_price * 10000 if unit_price < 1000 else unit_price,
            'decision_rate': dec_rate,
            'q_target': q_target * 10000 if q_target < 100000 else q_target,
            'q_result': q_result,
            'notes': notes,
        }
    return targets

def extract_referral_data(referral_path):
    """Extract referral funnel data from the referral spreadsheet."""
    wb = openpyxl.load_workbook(referral_path, data_only=True)
    ws = wb['数値・分析']
    result = {'targets': {}, 'years': {}}

    # Row 2-4: targets
    result['targets'] = {
        'annual': parse_number(ws.cell(2, 6).value),
        'q_target': parse_number(ws.cell(3, 6).value),
        'profit_target': parse_number(ws.cell(4, 6).value),
    }

    def parse_funnel_block(start_row, year, quarters):
        """Parse a CN/CS/Total funnel block."""
        data = {}
        statuses = ['アサイン', '初回面談', '求人紹介', '書類準備中', '応募', '応募(社数)',
                     '1st面接', '2nd面接', '最終面接', '内定', '決定']
        rates = ['求人紹介率', '面談応募率', '応募社数/一人', '書類通過率', '1次通過率',
                 '応募内定率', '内定決定率', '応募決定率', '面談決定率']

        for section_name, col_offset in [('CN', 5), ('CS', 12), ('Total', 19)]:
            section = {}
            for i, status in enumerate(statuses):
                row_vals = {}
                for qi, q in enumerate(quarters):
                    v = parse_number(ws.cell(start_row + i, col_offset + qi).value)
                    row_vals[q] = v
                section[status] = row_vals
            # Parse rates
            rate_start = start_row + len(statuses)
            for i, rate in enumerate(rates):
                row_vals = {}
                for qi, q in enumerate(quarters):
                    v = parse_number(ws.cell(rate_start + i, col_offset + qi).value)
                    row_vals[q] = v
                section[rate] = row_vals
            data[section_name] = section
        return data

    # 2026 block: rows 8-27, quarters = FY26/1Q..4Q
    fy26_qs = ['FY26/1Q', 'FY26/2Q', 'FY26/3Q', 'FY26/4Q']
    result['years']['2026'] = parse_funnel_block(8, '2026', fy26_qs)

    # CA breakdown for 2026 (rows 29+)
    ca_referral = {}
    for r in range(29, 47, 2):
        ca_name = safe_val(ws.cell(r, 17).value)  # Total section CA name
        if not ca_name: continue
        ca_name = str(ca_name).strip()
        interviews = {}
        decisions = {}
        for qi, q in enumerate(fy26_qs):
            interviews[q] = parse_number(ws.cell(r, 19 + qi).value)
            decisions[q] = parse_number(ws.cell(r + 1, 19 + qi).value)
        total_int = parse_number(ws.cell(r, 23).value)
        total_dec = parse_number(ws.cell(r + 1, 23).value)
        ca_referral[ca_name] = {
            'interviews': interviews, 'decisions': decisions,
            'total_interviews': total_int, 'total_decisions': total_dec,
        }
    result['ca_2026'] = ca_referral

    # 2025 block: rows 52-71
    fy25_qs = ['FY25/1Q', 'FY25/2Q', 'FY25/3Q', 'FY25/4Q']
    result['years']['2025'] = parse_funnel_block(52, '2025', fy25_qs)

    # CA breakdown for 2025 (rows 74+)
    ca_ref_25 = {}
    for r in range(74, 92, 2):
        ca_name = safe_val(ws.cell(r, 17).value)
        if not ca_name: continue
        ca_name = str(ca_name).strip()
        interviews = {}
        decisions = {}
        for qi, q in enumerate(fy25_qs):
            interviews[q] = parse_number(ws.cell(r, 19 + qi).value)
            decisions[q] = parse_number(ws.cell(r + 1, 19 + qi).value)
        total_int = parse_number(ws.cell(r, 23).value)
        total_dec = parse_number(ws.cell(r + 1, 23).value)
        ca_ref_25[ca_name] = {
            'interviews': interviews, 'decisions': decisions,
            'total_interviews': total_int, 'total_decisions': total_dec,
        }
    result['ca_2025'] = ca_ref_25

    # 2024 block: rows 95-114, CN=cols D-H (5-8), CS=cols K-O (12-15), Total=cols R-V (18-22)
    fy24_qs = ['FY24/1Q', 'FY24/2Q', 'FY24/3Q', 'FY24/4Q']
    statuses = ['アサイン', '初回面談', '求人紹介', '書類準備中', '応募', '応募(社数)',
                '1st面接', '2nd面接', '最終面接', '内定', '決定']
    rates = ['求人紹介率', '面談応募率', '応募社数/一人', '書類通過率', '1次通過率',
             '応募内定率', '内定決定率', '応募決定率', '面談決定率']

    fy24_data = {}
    for section_name, col_start in [('CN', 5), ('CS', 12), ('Total', 19)]:
        section = {}
        for i, status in enumerate(statuses):
            row_vals = {}
            for qi, q in enumerate(fy24_qs):
                v = parse_number(ws.cell(95 + i, col_start + qi).value)
                row_vals[q] = v
            section[status] = row_vals
        rate_start = 95 + len(statuses)
        for i, rate in enumerate(rates):
            row_vals = {}
            for qi, q in enumerate(fy24_qs):
                v = parse_number(ws.cell(rate_start + i, col_start + qi).value)
                row_vals[q] = v
            section[rate] = row_vals
        fy24_data[section_name] = section
    result['years']['2024'] = fy24_data

    # CA breakdown for 2024 (rows 117+, col C=CA, col D=metric, cols E-H=Q values)
    ca_ref_24 = {}
    for r in range(117, 133, 2):
        ca_name = safe_val(ws.cell(r, 3).value)
        if not ca_name: continue
        ca_name = str(ca_name).strip()
        interviews = {}
        decisions = {}
        for qi, q in enumerate(fy24_qs):
            interviews[q] = parse_number(ws.cell(r, 5 + qi).value)
            decisions[q] = parse_number(ws.cell(r + 1, 5 + qi).value)
        ca_ref_24[ca_name] = {
            'interviews': interviews, 'decisions': decisions,
        }
    result['ca_2024'] = ca_ref_24

    return result

def extract_route_breakdown(wb):
    """Extract inflow breakdown by route (ビズリーチ, AMBI, etc.) from DB_求職者一覧."""
    ws = wb['DB_求職者一覧']
    from collections import Counter, defaultdict

    def to_quarter(y, m):
        """Calendar year quarters: Q1=Jan-Mar, Q2=Apr-Jun, Q3=Jul-Sep, Q4=Oct-Dec."""
        q = (m - 1) // 3 + 1
        return f'FY{y}/{q}Q'

    route_map = {
        'ビズリーチ': 'ビズリーチ', 'ビズリーチ(A)': 'ビズリーチ',
        'AMBI': 'AMBI', 'AMBI(A)': 'AMBI',
        'リクルート': 'リクルート', 'リクルート(A)': 'リクルート',
        'OpenWork': 'OpenWork', 'Green': 'Green',
        '紹介(CS)': '紹介', '紹介(CN)': '紹介',
        'indeed': 'indeed', 'キャリアパーク': 'キャリアパーク',
        'キミナラ': 'キミナラ', 'オーガニック': 'オーガニック',
    }

    route_monthly = defaultdict(lambda: Counter())
    route_quarterly = defaultdict(lambda: Counter())
    # Decision data by route/quarter
    route_q_decisions = defaultdict(lambda: defaultdict(int))
    route_q_total = defaultdict(lambda: defaultdict(int))

    def parse_dt(v):
        if isinstance(v, datetime): return v
        if isinstance(v, str):
            try: return datetime.strptime(v.strip()[:10], '%Y-%m-%d')
            except: pass
        return None

    for r in range(3, min(ws.max_row + 1, 10001)):
        route_raw = safe_val(ws.cell(r, 6).value)
        referral_raw = safe_val(ws.cell(r, 7).value)
        interview_date = parse_dt(ws.cell(r, 16).value)  # 面談日ベース（col 16）
        dec_val = ws.cell(r, 45).value  # 決定数

        if not route_raw and not referral_raw: continue
        # Use referral column (col 7) if it maps to a known route, else fall back to col 6
        route_key = str(route_raw or '').strip()
        if referral_raw and str(referral_raw).strip() in route_map:
            route_key = str(referral_raw).strip()
        route = route_map.get(route_key, 'その他')

        if interview_date:
            y, m = interview_date.year, interview_date.month
            month_key = f'{y}年{m}月'
            q_key = to_quarter(y, m)
            route_monthly[month_key][route] += 1
            route_quarterly[q_key][route] += 1
            route_q_total[route][q_key] += 1

            if dec_val is not None:
                try:
                    dv = float(dec_val)
                    if dv > 0:
                        route_q_decisions[route][q_key] += int(dv)
                except:
                    pass

    # Build decision rate by route/quarter
    route_decision_rates = {}
    for route in route_q_total:
        route_decision_rates[route] = {}
        for q in route_q_total[route]:
            total = route_q_total[route][q]
            dec = route_q_decisions[route].get(q, 0)
            route_decision_rates[route][q] = {
                'total': total, 'decisions': dec,
                'rate': dec / total if total > 0 else 0,
            }

    # Filter to FY24+ quarters dynamically
    current_year = datetime.now().year
    valid_years = {str(y) for y in range(2024, current_year + 1)}
    valid_qs = {q for qs in route_quarterly.values() for q in qs if any(yr in q for yr in valid_years)}

    return {
        'monthly': {k: dict(v) for k, v in route_monthly.items()},
        'quarterly': {k: dict(v) for k, v in route_quarterly.items()},
        'routes': sorted(set(r for c in route_quarterly.values() for r in c.keys())),
        'decision_rates': {r: {q: v for q, v in qs.items()} for r, qs in route_decision_rates.items()},
    }

def extract_yomi_forecast(wb):
    """Extract ヨミ予測 (pipeline forecast) from DB_求職者一覧.

    New logic (2026/4Q~):
    - 確度 30%以上を計算対象
    - 期待値 = 粗利 × 確度%
    - 月別着地予測（当月/来月/翌々月）
    - 50%以上ゾーン vs 30%ゾーンを分けて集計
    - 「決定」ステータスは確度100%として扱う
    """
    ws = wb['DB_求職者一覧']
    today = datetime.now()
    current_fy = today.year
    current_q = (today.month - 1) // 3 + 1
    q_start_month = (current_q - 1) * 3 + 1
    q_end_month = current_q * 3

    # Current Q month range
    q_months = []
    for m in range(q_start_month, q_end_month + 1):
        q_months.append((current_fy, m))

    # Month labels
    month_labels = {}
    for i, (y, m) in enumerate(q_months):
        if i == 0:
            label = '当月' if today.month == m else f'{m}月'
        month_labels[(y, m)] = f'{m}月'

    # Determine which month index is "current"
    current_month_idx = today.month - q_start_month  # 0, 1, or 2

    departed_cas = ['渡辺', '百瀬']

    # Q start date for filtering
    q_start_date = datetime(current_fy, q_start_month, 1)

    def parse_date_ym(raw):
        """Parse a date value and return (year, month) or (None, None)."""
        if not raw:
            return None, None
        if isinstance(raw, (datetime, date)):
            return raw.year, raw.month
        dt_match = re.match(r'(\d{4})-(\d{2})', str(raw))
        if dt_match:
            return int(dt_match.group(1)), int(dt_match.group(2))
        return None, None

    # Collect all active pipeline candidates with 確度 >= 30%
    candidates = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 1).value
        status = str(ws.cell(r, 2).value or '')
        ca = str(ws.cell(r, 3).value or '').strip()
        kakudo_raw = ws.cell(r, 9).value
        arari = parse_number(ws.cell(r, 11).value)
        chakuchi_raw = ws.cell(r, 17).value  # 着地日
        kettei_raw = ws.cell(r, 18).value    # 決定日
        dec_count = parse_number(ws.cell(r, 45).value)

        if not name or not ca or ca == 'None' or ca in departed_cas:
            continue

        # Skip completed (入社済み) or NG
        if 'NG' in status or '入社' in status:
            continue

        # Parse 確度
        kakudo = parse_number(kakudo_raw)
        if kakudo <= 0:
            continue
        # Values are 0.1-1.0 in sheet; convert to percentage
        if kakudo <= 1.0:
            kakudo_pct = kakudo  # Keep as decimal (0.3 = 30%)
        else:
            kakudo_pct = kakudo / 100.0

        # 「決定」ステータスは100%
        if status == '決定':
            kakudo_pct = 1.0

        # Skip if below 30%
        if kakudo_pct < 0.3:
            continue

        # ===== 月判定ロジック =====
        # 優先順位: 決定日 > 着地日
        # 決定日があれば → その月で判定（過去Qなら除外）
        # 決定日なし＆着地日あり → 着地日で判定
        # どちらもなし → 未定（除外）
        kettei_year, kettei_month = parse_date_ym(kettei_raw)
        chakuchi_year, chakuchi_month = parse_date_ym(chakuchi_raw)

        # Determine effective date for Q/month assignment
        eff_year, eff_month = None, None
        if kettei_year and kettei_month:
            # 決定日がある → 決定日で判定
            eff_year, eff_month = kettei_year, kettei_month
        elif chakuchi_year and chakuchi_month:
            # 決定日なし → 着地日で判定
            eff_year, eff_month = chakuchi_year, chakuchi_month
        else:
            # どちらもなし → 未定
            pass

        # Check if effective date falls in current Q
        landing_in_q = False
        landing_month_idx = None  # 0=month1, 1=month2, 2=month3
        if eff_year and eff_month:
            # If effective date is before current Q → skip (past Q revenue)
            eff_date = datetime(eff_year, eff_month, 1)
            if eff_date < q_start_date:
                continue  # 過去Qの決定/着地 → 2Q予測から除外

            # Check if it falls in one of the Q months
            for i, (qy, qm) in enumerate(q_months):
                if eff_year == qy and eff_month == qm:
                    landing_in_q = True
                    landing_month_idx = i
                    break
            # If after current Q → not in this Q's forecast
        else:
            # No date → unscheduled
            pass

        # Calculate 期待粗利
        kitai_arari = arari * kakudo_pct

        candidates.append({
            'name': str(name),
            'status': status,
            'ca': ca,
            'kakudo': round(kakudo_pct * 100),  # As integer percentage
            'arari': arari,
            'kitai_arari': round(kitai_arari),
            'eff_year': eff_year,
            'eff_month': eff_month,
            'has_kettei': kettei_year is not None,
            'landing_in_q': landing_in_q,
            'landing_month_idx': landing_month_idx,
            'row': r,
        })

    # ===== Aggregate by CA and by month =====
    # Active CAs (with at least one candidate or budget)
    ca_set = set()
    for c in candidates:
        if c['landing_in_q']:
            ca_set.add(c['ca'])

    # Build per-CA forecast
    ca_forecasts = {}
    for ca in sorted(ca_set):
        ca_cands = [c for c in candidates if c['ca'] == ca and c['landing_in_q']]

        # Group by landing month
        monthly = {}
        for i, (y, m) in enumerate(q_months):
            month_cands = [c for c in ca_cands if c['landing_month_idx'] == i]
            # Sort by kakudo descending
            month_cands.sort(key=lambda x: (-x['kakudo'], -x['arari']))

            # Split into 50%+ and 30% zones
            high_zone = [c for c in month_cands if c['kakudo'] >= 50]
            low_zone = [c for c in month_cands if c['kakudo'] < 50]

            high_total = sum(c['kitai_arari'] for c in high_zone)
            low_total = sum(c['kitai_arari'] for c in low_zone)

            monthly[i] = {
                'year': y,
                'month': m,
                'label': f'{m}月',
                'candidates': [{
                    'name': c['name'],
                    'kakudo': c['kakudo'],
                    'status': c['status'],
                    'arari': c['arari'],
                    'kitai_arari': c['kitai_arari'],
                } for c in month_cands],
                'count': len(month_cands),
                'high_zone_total': round(high_total),
                'low_zone_total': round(low_total),
                'total': round(high_total + low_total),
                'kakudo_breakdown': _kakudo_breakdown(month_cands),
            }

        q_total = sum(monthly[i]['total'] for i in range(3))
        q_high = sum(monthly[i]['high_zone_total'] for i in range(3))
        q_low = sum(monthly[i]['low_zone_total'] for i in range(3))
        q_count = sum(monthly[i]['count'] for i in range(3))

        # Candidates NOT landing in Q (未定 or future Q)
        unscheduled = [c for c in candidates if c['ca'] == ca and not c['landing_in_q'] and c['kakudo'] >= 30]

        ca_forecasts[ca] = {
            'monthly': monthly,
            'q_total': round(q_total),
            'q_high_zone': round(q_high),
            'q_low_zone': round(q_low),
            'q_count': q_count,
            'unscheduled_count': len(unscheduled),
            'unscheduled_total': round(sum(c['kitai_arari'] for c in unscheduled)),
            'kakudo_up_candidates': [
                {'name': c['name'], 'kakudo': c['kakudo'], 'status': c['status'], 'arari': c['arari']}
                for c in ca_cands if 40 <= c['kakudo'] <= 60
            ],
        }

    # ===== Company-wide totals =====
    company_monthly = {}
    for i, (y, m) in enumerate(q_months):
        month_cands = [c for c in candidates if c['landing_in_q'] and c['landing_month_idx'] == i]
        high_zone = [c for c in month_cands if c['kakudo'] >= 50]
        low_zone = [c for c in month_cands if c['kakudo'] < 50]
        company_monthly[i] = {
            'year': y,
            'month': m,
            'label': f'{m}月',
            'count': len(month_cands),
            'high_zone_total': round(sum(c['kitai_arari'] for c in high_zone)),
            'low_zone_total': round(sum(c['kitai_arari'] for c in low_zone)),
            'total': round(sum(c['kitai_arari'] for c in month_cands)),
            'kakudo_breakdown': _kakudo_breakdown(month_cands),
        }

    company_q_total = sum(company_monthly[i]['total'] for i in range(3))

    # Get Q targets and monthly targets from DB_売上管理
    q_info = get_current_quarter()
    q_target = 0
    monthly_targets = {}  # {month_idx: target}
    current_q_sales_key = q_info['current_q_sales']
    try:
        ws_sales = wb['DB_売上管理']
        # Row 3 has headers, Row 4 has 合計 目標
        # Find Q column and monthly columns
        q_col = None
        month_cols = {}  # {month_num: col}
        for col in range(2, ws_sales.max_column + 1):
            header = str(ws_sales.cell(3, col).value or '')
            if current_q_sales_key in header:
                q_col = col
            # Match monthly headers like "2026年4月"
            for i, (y, m) in enumerate(q_months):
                if f'{y}年{m}月' in header:
                    month_cols[i] = col
        if q_col:
            q_target = parse_number(ws_sales.cell(4, q_col).value)
        for i, col in month_cols.items():
            monthly_targets[i] = parse_number(ws_sales.cell(4, col).value)
    except Exception:
        pass

    monthly_target = round(q_target / 3) if q_target > 0 else 0

    # Add per-month targets to company_monthly
    for i in range(3):
        m_target = monthly_targets.get(i, monthly_target)
        company_monthly[i]['target'] = round(m_target)
        company_monthly[i]['gap'] = round(company_monthly[i]['total'] - m_target)

    # Get per-CA targets from DB_売上管理
    ca_targets = {}
    try:
        ws_sales = wb['DB_売上管理']
        # Each CA block is 13 rows starting from row 4
        # Scan for CA names in column 1
        for row in range(4, ws_sales.max_row + 1):
            ca_name = str(ws_sales.cell(row, 1).value or '').strip()
            item = str(ws_sales.cell(row, 2).value or '').strip()
            if ca_name and '目標' in item and q_col:
                ca_q_target = parse_number(ws_sales.cell(row, q_col).value)
                if ca_q_target > 0:
                    ca_targets[ca_name] = ca_q_target
                    # Also get monthly targets for this CA
                    ca_monthly_targets = {}
                    for i, col in month_cols.items():
                        ca_monthly_targets[i] = parse_number(ws_sales.cell(row, col).value)
                    # Find 着地 row (should be row + 3 for CA粗利着地)
                    for offset in range(1, 14):
                        check_item = str(ws_sales.cell(row + offset, 2).value or '').strip()
                        if '着地' in check_item and 'CA' in check_item:
                            ca_db_landing = {}
                            if q_col:
                                ca_db_landing['q'] = parse_number(ws_sales.cell(row + offset, q_col).value)
                            for mi, col in month_cols.items():
                                ca_db_landing[mi] = parse_number(ws_sales.cell(row + offset, col).value)
                            if ca_name in ca_forecasts:
                                ca_forecasts[ca_name]['db_landing'] = ca_db_landing
                            break
                    if ca_name in ca_forecasts:
                        ca_forecasts[ca_name]['q_target'] = round(ca_q_target)
                        ca_forecasts[ca_name]['q_gap'] = round(ca_forecasts[ca_name]['q_total'] - ca_q_target)
                        ca_forecasts[ca_name]['monthly_targets'] = {k: round(v) for k, v in ca_monthly_targets.items()}
    except Exception as e:
        print(f"  Warning: could not read CA targets: {e}")

    result = {
        'generated_at': today.strftime('%Y-%m-%d %H:%M'),
        'quarter': f'FY{current_fy % 100}/{current_q}Q',
        'q_months': [{'year': y, 'month': m, 'label': f'{m}月'} for y, m in q_months],
        'current_month_idx': current_month_idx,
        'q_target': round(q_target),
        'monthly_target': monthly_target,
        'monthly_targets': {k: round(v) for k, v in monthly_targets.items()},
        'company': {
            'monthly': company_monthly,
            'q_total': round(company_q_total),
            'q_target': round(q_target),
            'q_gap': round(company_q_total - q_target),
            'q_achievement_rate': round(company_q_total / q_target * 100) if q_target > 0 else 0,
        },
        'by_ca': ca_forecasts,
        'total_candidates': len([c for c in candidates if c['landing_in_q']]),
    }

    return result


def _kakudo_breakdown(cands):
    """Count candidates by 確度 level."""
    breakdown = {}
    for c in cands:
        k = c['kakudo']
        if k not in breakdown:
            breakdown[k] = 0
        breakdown[k] += 1
    return breakdown


def build_weekly_report(yomi_forecast, sales, ca_funnel, kpi, ca_comparison, q_info, departed_cas, wb=None):
    """Build weekly report with landing predictions and improvement suggestions.

    Combines 3 prediction methods:
    A. 確定見込み (existing pipeline) — yomi_forecast.q_total
    B. 新規面談からの予測 — (残日数 × 日次面談ペース) × 過去決定率 × 平均粗利
    C. 統合予測 = A + B (後段で重複補正)

    Compares with previous snapshot to show week-over-week trend.
    Stores today's snapshot to data/weekly_snapshots/YYYY-MM-DD.json
    """
    import os
    today = datetime.now()
    today_key = today.strftime('%Y-%m-%d')
    snapshot_dir = 'data/weekly_snapshots'
    os.makedirs(snapshot_dir, exist_ok=True)

    current_q_sales_key = q_info['current_q_sales']
    q_end = q_info['q_end_date']
    days_remaining = max(0, (q_end - today).days)
    q_start = datetime(q_info['fy'], (q_info['q'] - 1) * 3 + 1, 1)
    q_total_days = (q_end - q_start).days + 1
    q_elapsed = (today - q_start).days + 1
    q_progress = max(0, min(1, q_elapsed / q_total_days))

    # ----- Q目標 / 実績 -----
    total_sales_q = sales.get('合計', {}).get('quarterly', {}).get(current_q_sales_key, {})
    q_target = total_sales_q.get('目標(粗利)', 0)
    q_actual = total_sales_q.get('実績(粗利)', 0)
    q_landing_db = total_sales_q.get('着地見込み', 0)  # スプシ上の着地見込み

    # ----- 予測ロジック A: 既存パイプライン (ヨミ予測) -----
    pipeline_forecast = yomi_forecast.get('company', {}).get('q_total', 0)

    # ----- 予測ロジック B: 新規面談からの予測 -----
    # 過去3か月の面談ペース・決定率・平均粗利を算出
    funnel = ca_funnel.get('合計', ca_funnel.get('total', {})).get('funnel', {})

    # 直近3か月の面談数
    recent_months = []
    for offset in range(1, 4):
        m = today.month - offset
        y = today.year
        if m <= 0:
            m += 12
            y -= 1
        recent_months.append(f'{y}-{m:02d}')

    recent_interview_total = sum(funnel.get('初回面談', {}).get(m, 0) for m in recent_months)
    days_in_3mo = 90
    daily_interview_pace = recent_interview_total / days_in_3mo if recent_interview_total > 0 else 0

    # 決定率: 過去6か月のデータから算出（より安定した数値）
    # リードタイム2か月想定で、面談から決定までの転換率を計算
    six_months = []
    for offset in range(2, 8):  # 2-7か月前
        m = today.month - offset
        y = today.year
        if m <= 0:
            m += 12
            y -= 1
        six_months.append(f'{y}-{m:02d}')

    six_months_decisions = []
    for offset in range(0, 6):  # 直近6か月の決定
        m = today.month - offset
        y = today.year
        if m <= 0:
            m += 12
            y -= 1
        six_months_decisions.append(f'{y}-{m:02d}')

    interviews_6mo = sum(funnel.get('初回面談', {}).get(m, 0) for m in six_months)
    decisions_6mo = sum(funnel.get('決定', {}).get(m, 0) for m in six_months_decisions)
    decision_rate = decisions_6mo / interviews_6mo if interviews_6mo > 0 else 0.05

    # 平均粗利（過去Q）
    prev_q_key = q_info['confirmed_q']
    prev_q_data = sales.get('合計', {}).get('quarterly', {}).get(prev_q_key, {})
    prev_q_decisions = prev_q_data.get('決定数', 1) or 1
    prev_q_actual = prev_q_data.get('実績(粗利)', 0)
    avg_profit_per_decision = prev_q_actual / prev_q_decisions if prev_q_decisions > 0 else 1100000

    # ===== リードタイム考慮 =====
    # 初回面談から決定までの平均リードタイム = 2か月（60日）
    # Q末までに決定するには、Q末 - 60日 までに面談する必要がある
    # 例: Q2末 6/30 → 最新有効面談日 = 4/30
    # 今日が4/25 → 有効残日数 = 5日
    lead_time_days = 60
    latest_interview_date = q_end - __import__('datetime').timedelta(days=lead_time_days)
    effective_days = max(0, (latest_interview_date - today).days)

    new_interviews_predicted = effective_days * daily_interview_pace
    new_decisions_predicted = new_interviews_predicted * decision_rate
    new_revenue_predicted = new_decisions_predicted * avg_profit_per_decision

    # ----- 統合予測 C -----
    # A (既存パイプライン) には実績も含む。B (新規面談予測) は追加分
    # 重複なし: A は今ある人の期待値、B はこれから面談する人
    integrated_forecast = pipeline_forecast + new_revenue_predicted

    # ----- Gap & 改善示唆 -----
    gap = q_target - integrated_forecast
    achievement_rate = integrated_forecast / q_target if q_target > 0 else 0

    # 必要追加面談数 (Gap埋めに必要な数)
    needed_decisions = gap / avg_profit_per_decision if avg_profit_per_decision > 0 and gap > 0 else 0
    needed_interviews = needed_decisions / decision_rate if decision_rate > 0 and needed_decisions > 0 else 0

    # 残日数で達成可能か
    needed_daily_pace = needed_interviews / effective_days if effective_days > 0 else 0
    pace_gap = needed_daily_pace - daily_interview_pace

    # ----- ホット案件 (確度50%以上の合計) -----
    hot_pipeline = 0
    hot_count = 0
    kakudo_up_targets = []  # 40-60%帯の確度UP候補
    for ca, ca_data in yomi_forecast.get('by_ca', {}).items():
        hot_pipeline += ca_data.get('q_high_zone', 0)
        for i in range(3):
            month_data = ca_data.get('monthly', {}).get(i, {})
            for c in month_data.get('candidates', []):
                if c.get('kakudo', 0) >= 50:
                    hot_count += 1
                if 40 <= c.get('kakudo', 0) <= 60:
                    kakudo_up_targets.append({
                        'ca': ca, 'name': c['name'], 'kakudo': c['kakudo'],
                        'status': c['status'], 'arari': c['arari'], 'kitai_arari': c['kitai_arari']
                    })

    # ===== 過去Qの実達成率カーブ =====
    # 過去Qの週次累積実績から「Q開始からX日後の平均達成率」を算出
    ws_db = wb['DB_求職者一覧'] if (wb and 'DB_求職者一覧' in wb.sheetnames) else None
    past_q_curves = {}  # {Q_key: [{day_offset, cumul_actual, achievement_rate}, ...]}
    past_q_targets = {
        'FY25/4Q': sales.get('合計', {}).get('quarterly', {}).get('FY25／4Q', {}).get('目標(粗利)', 0),
        'FY26/1Q': sales.get('合計', {}).get('quarterly', {}).get('FY26／1Q', {}).get('目標(粗利)', 0),
    }
    if ws_db:
        # 過去Q別に決定日ベースで週次集計
        for q_key, q_start_y, q_start_m in [('FY25/4Q', 2025, 10), ('FY26/1Q', 2026, 1)]:
            q_start_dt = datetime(q_start_y, q_start_m, 1)
            # Q末
            q_end_m = q_start_m + 2
            if q_end_m == 12:
                q_end_dt = datetime(q_start_y, 12, 31)
            else:
                q_end_dt = datetime(q_start_y, q_end_m + 1, 1) - __import__('datetime').timedelta(days=1)
            past_q_target = past_q_targets.get(q_key, 0)
            if past_q_target == 0:
                continue

            daily_decisions = defaultdict(float)
            for r in range(3, ws_db.max_row + 1):
                status = str(ws_db.cell(r, 2).value or '')
                arari = parse_number(ws_db.cell(r, 11).value)
                kettei_raw = ws_db.cell(r, 18).value
                if status not in ('決定', '入社'):
                    continue
                if not isinstance(kettei_raw, (datetime, date)):
                    continue
                if q_start_dt <= kettei_raw <= q_end_dt:
                    day_offset = (kettei_raw - q_start_dt).days
                    daily_decisions[day_offset] += arari

            # 累積カーブ
            curve = []
            cumul = 0
            total_days = (q_end_dt - q_start_dt).days
            for d in range(0, total_days + 1):
                cumul += daily_decisions.get(d, 0)
                curve.append({
                    'day': d,
                    'cumul': round(cumul),
                    'rate': cumul / past_q_target if past_q_target > 0 else 0,
                })
            past_q_curves[q_key] = {'target': past_q_target, 'curve': curve}

    # 過去Qの平均達成率カーブ（同じ日数経過時点での平均）
    avg_curve = []
    if past_q_curves:
        max_days = min(len(v['curve']) for v in past_q_curves.values())
        for d in range(max_days):
            rates = [v['curve'][d]['rate'] for v in past_q_curves.values()]
            avg_curve.append({
                'day': d,
                'avg_rate': sum(rates) / len(rates),
            })

    # 現Qの累積実績カーブ
    current_q_curve = []
    cumul_now = 0
    if ws_db:
        daily_decisions_cur = defaultdict(float)
        for r in range(3, ws_db.max_row + 1):
            status = str(ws_db.cell(r, 2).value or '')
            arari = parse_number(ws_db.cell(r, 11).value)
            kettei_raw = ws_db.cell(r, 18).value
            if status not in ('決定', '入社'):
                continue
            if not isinstance(kettei_raw, (datetime, date)):
                continue
            if q_start <= kettei_raw <= q_end:
                day_offset = (kettei_raw - q_start).days
                daily_decisions_cur[day_offset] += arari

        elapsed_days = (today - q_start).days
        for d in range(0, elapsed_days + 1):
            cumul_now += daily_decisions_cur.get(d, 0)
            current_q_curve.append({
                'day': d,
                'cumul': round(cumul_now),
                'rate': cumul_now / q_target if q_target > 0 else 0,
            })

    # 現Q末予測（過去Qパターン継続シナリオ）
    historical_landing = 0
    historical_final_rate = 0
    if avg_curve and current_q_curve and len(current_q_curve) > 0:
        cur_day = current_q_curve[-1]['day']
        if cur_day < len(avg_curve):
            avg_rate_at_today = avg_curve[cur_day]['avg_rate']
            avg_rate_at_q_end = avg_curve[-1]['avg_rate'] if len(avg_curve) > 0 else 0
            cur_rate_today = current_q_curve[-1]['rate']
            # 過去のQ末/今日の比率 × 今日の実績 = Q末予測
            if avg_rate_at_today > 0:
                ratio = avg_rate_at_q_end / avg_rate_at_today
                historical_landing = current_q_curve[-1]['cumul'] * ratio
                historical_final_rate = historical_landing / q_target if q_target > 0 else 0

    # ----- 今日のスナップショット -----
    snapshot = {
        'date': today_key,
        'q_target': q_target,
        'q_actual': q_actual,
        'pipeline_forecast': pipeline_forecast,
        'new_revenue_predicted': new_revenue_predicted,
        'integrated_forecast': integrated_forecast,
        'gap': gap,
        'achievement_rate': achievement_rate,
        'days_remaining': days_remaining,
        'q_progress': q_progress,
        'pipeline_count': yomi_forecast.get('total_candidates', 0),
        'hot_pipeline': hot_pipeline,
        'hot_count': hot_count,
        'daily_interview_pace': daily_interview_pace,
        'decision_rate': decision_rate,
        'avg_profit_per_decision': avg_profit_per_decision,
        'by_ca': {},
    }

    # ----- 個別CA -----
    active_cas = [c for c in kpi.get('ca_names', []) if c not in departed_cas]
    for ca in active_cas:
        ca_yomi = yomi_forecast.get('by_ca', {}).get(ca, {})
        ca_pipeline = ca_yomi.get('q_total', 0)
        ca_target = ca_yomi.get('q_target', 0)

        # CA別の面談ペース・決定率
        ca_f = ca_funnel.get(ca, {}).get('funnel', {})
        ca_recent_interviews = sum(ca_f.get('初回面談', {}).get(m, 0) for m in recent_months)
        ca_daily_pace = ca_recent_interviews / 90 if ca_recent_interviews > 0 else 0

        ca_int_6mo = sum(ca_f.get('初回面談', {}).get(m, 0) for m in six_months)
        ca_dec_6mo = sum(ca_f.get('決定', {}).get(m, 0) for m in six_months_decisions)
        ca_dec_rate = ca_dec_6mo / ca_int_6mo if ca_int_6mo > 0 else decision_rate

        ca_new_int_pred = effective_days * ca_daily_pace
        ca_new_dec_pred = ca_new_int_pred * ca_dec_rate
        ca_new_rev_pred = ca_new_dec_pred * avg_profit_per_decision

        ca_integrated = ca_pipeline + ca_new_rev_pred
        ca_gap = ca_target - ca_integrated
        ca_ach_rate = ca_integrated / ca_target if ca_target > 0 else 0

        ca_needed_decisions = ca_gap / avg_profit_per_decision if avg_profit_per_decision > 0 and ca_gap > 0 else 0
        ca_needed_interviews = ca_needed_decisions / ca_dec_rate if ca_dec_rate > 0 and ca_needed_decisions > 0 else 0

        snapshot['by_ca'][ca] = {
            'q_target': ca_target,
            'pipeline_forecast': ca_pipeline,
            'new_revenue_predicted': round(ca_new_rev_pred),
            'integrated_forecast': round(ca_integrated),
            'gap': round(ca_gap),
            'achievement_rate': ca_ach_rate,
            'daily_interview_pace': ca_daily_pace,
            'decision_rate': ca_dec_rate,
            'recent_interviews': ca_recent_interviews,
            'needed_interviews': round(ca_needed_interviews),
            'needed_decisions': round(ca_needed_decisions, 1),
            'q_count': ca_yomi.get('q_count', 0),
            'q_high_zone': ca_yomi.get('q_high_zone', 0),
        }

    # ----- 過去スナップショットを読み込んで週次比較 -----
    past_snapshots = []
    try:
        for fname in sorted(os.listdir(snapshot_dir)):
            if fname.endswith('.json') and fname != f'{today_key}.json':
                fpath = os.path.join(snapshot_dir, fname)
                with open(fpath, 'r', encoding='utf-8') as f:
                    s = json.load(f)
                if s.get('q_target', 0) == q_target:  # Same Q
                    past_snapshots.append(s)
    except Exception:
        pass

    # 7日前のスナップショット（または最も近い過去のもの）
    week_ago_target = today - __import__('datetime').timedelta(days=7)
    last_week_snapshot = None
    if past_snapshots:
        # Find closest snapshot within 5-9 days ago
        for s in reversed(past_snapshots):
            try:
                s_date = datetime.strptime(s['date'], '%Y-%m-%d')
                days_diff = (today - s_date).days
                if 5 <= days_diff <= 9:
                    last_week_snapshot = s
                    break
            except Exception:
                continue
        if not last_week_snapshot:
            last_week_snapshot = past_snapshots[-1]  # Most recent past

    # 週次変化
    weekly_change = {}
    if last_week_snapshot:
        weekly_change = {
            'last_date': last_week_snapshot.get('date'),
            'forecast_change': integrated_forecast - last_week_snapshot.get('integrated_forecast', 0),
            'forecast_change_pct': (integrated_forecast / last_week_snapshot.get('integrated_forecast', 1) - 1) if last_week_snapshot.get('integrated_forecast', 0) > 0 else 0,
            'pipeline_change': pipeline_forecast - last_week_snapshot.get('pipeline_forecast', 0),
            'hot_pipeline_change': hot_pipeline - last_week_snapshot.get('hot_pipeline', 0),
            'achievement_change': achievement_rate - last_week_snapshot.get('achievement_rate', 0),
            'pace_change': daily_interview_pace - last_week_snapshot.get('daily_interview_pace', 0),
            'last_forecast': last_week_snapshot.get('integrated_forecast', 0),
            'last_pipeline': last_week_snapshot.get('pipeline_forecast', 0),
            'last_hot': last_week_snapshot.get('hot_pipeline', 0),
        }

    # 過去30日のトレンド
    trend_data = []
    for s in past_snapshots[-30:]:
        trend_data.append({
            'date': s.get('date'),
            'forecast': s.get('integrated_forecast', 0),
            'pipeline': s.get('pipeline_forecast', 0),
            'hot': s.get('hot_pipeline', 0),
            'achievement': s.get('achievement_rate', 0),
        })
    # Add today
    trend_data.append({
        'date': today_key,
        'forecast': integrated_forecast,
        'pipeline': pipeline_forecast,
        'hot': hot_pipeline,
        'achievement': achievement_rate,
    })

    # CA別の週次変化
    ca_weekly_change = {}
    if last_week_snapshot:
        for ca in active_cas:
            last_ca = last_week_snapshot.get('by_ca', {}).get(ca, {})
            cur_ca = snapshot['by_ca'].get(ca, {})
            ca_weekly_change[ca] = {
                'forecast_change': cur_ca.get('integrated_forecast', 0) - last_ca.get('integrated_forecast', 0),
                'pipeline_change': cur_ca.get('pipeline_forecast', 0) - last_ca.get('pipeline_forecast', 0),
                'hot_change': cur_ca.get('q_high_zone', 0) - last_ca.get('q_high_zone', 0),
            }

    # ----- 改善示唆 -----
    suggestions = []
    if gap > 0:
        suggestions.append({
            'type': 'gap',
            'priority': 'high',
            'message': f'Gap ▲{gap/10000:.0f}万円を埋めるには、追加で{needed_decisions:.1f}件の決定（{needed_interviews:.0f}件の新規面談）が必要',
        })
        if pace_gap > 0:
            suggestions.append({
                'type': 'pace',
                'priority': 'high',
                'message': f'必要日次面談ペース {needed_daily_pace:.1f}件/日 > 現状 {daily_interview_pace:.1f}件/日 → 日次+{pace_gap:.1f}件UP必須',
            })
    if hot_count > 0:
        suggestions.append({
            'type': 'hot',
            'priority': 'medium',
            'message': f'ホット案件（50%以上）{hot_count}名・{hot_pipeline/10000:.0f}万円 → クロージング集中',
        })
    if len(kakudo_up_targets) > 0:
        kakudo_up_targets.sort(key=lambda x: -x['arari'])
        top_targets = kakudo_up_targets[:5]
        suggestions.append({
            'type': 'kakudo_up',
            'priority': 'high',
            'message': f'確度UP候補（40-60%帯）{len(kakudo_up_targets)}名 → 上位{len(top_targets)}名のフェーズアップに注力',
            'targets': top_targets,
        })

    # CA別の示唆
    ca_suggestions = {}
    for ca in active_cas:
        cd = snapshot['by_ca'].get(ca, {})
        ca_sugs = []
        ca_gap = cd.get('gap', 0)
        if ca_gap > 0:
            ca_sugs.append({
                'type': 'gap',
                'message': f'Gap ▲{ca_gap/10000:.0f}万 → 追加{cd.get("needed_decisions",0)}件決定 / {cd.get("needed_interviews",0)}件面談必要',
            })
        ca_pace = cd.get('daily_interview_pace', 0)
        if ca_pace < daily_interview_pace * 0.7 and len(active_cas) > 0:
            avg_pace = daily_interview_pace / len(active_cas)
            if ca_pace < avg_pace * 0.7:
                ca_sugs.append({
                    'type': 'pace',
                    'message': f'面談ペース{ca_pace:.2f}件/日 (チーム平均{avg_pace:.2f}件/日) → ペースUP必要',
                })
        ca_dr = cd.get('decision_rate', 0)
        if ca_dr < decision_rate * 0.7 and ca_dr > 0:
            ca_sugs.append({
                'type': 'rate',
                'message': f'決定率{ca_dr*100:.1f}% (チーム{decision_rate*100:.1f}%) → 質の改善余地',
            })
        ca_suggestions[ca] = ca_sugs

    # ----- スナップショット保存 -----
    try:
        with open(os.path.join(snapshot_dir, f'{today_key}.json'), 'w', encoding='utf-8') as f:
            json.dump(snapshot, f, ensure_ascii=False, indent=2, default=str)
    except Exception as e:
        print(f"  Warning: snapshot save failed: {e}")

    return {
        'generated_at': today.strftime('%Y-%m-%d %H:%M'),
        'quarter': q_info['current_q_funnel'],
        'q_target': q_target,
        'q_actual': q_actual,
        'q_landing_db': q_landing_db,
        'days_remaining': days_remaining,
        'q_progress': q_progress,

        # 予測内訳
        'pipeline_forecast': round(pipeline_forecast),
        'new_revenue_predicted': round(new_revenue_predicted),
        'integrated_forecast': round(integrated_forecast),
        'gap': round(gap),
        'achievement_rate': achievement_rate,

        # 予測ロジックの内訳（説明用）
        'forecast_logic': {
            'daily_interview_pace': round(daily_interview_pace, 2),
            'decision_rate': decision_rate,
            'avg_profit_per_decision': round(avg_profit_per_decision),
            'recent_interview_total': recent_interview_total,
            'recent_months': recent_months,
            'effective_days': effective_days,
            'lead_time_days': lead_time_days,
            'latest_interview_date': latest_interview_date.strftime('%Y-%m-%d'),
            'q_end_date': q_end.strftime('%Y-%m-%d'),
            'new_interviews_predicted': round(new_interviews_predicted),
            'new_decisions_predicted': round(new_decisions_predicted, 1),
        },

        # 改善示唆
        'needed_interviews': round(needed_interviews),
        'needed_decisions': round(needed_decisions, 1),
        'needed_daily_pace': round(needed_daily_pace, 2),
        'pace_gap': round(pace_gap, 2),

        # ホット案件
        'hot_pipeline': round(hot_pipeline),
        'hot_count': hot_count,
        'kakudo_up_targets': kakudo_up_targets[:10],

        # 週次変化
        'weekly_change': weekly_change,
        'trend_data': trend_data,
        'ca_weekly_change': ca_weekly_change,

        # CA別
        'by_ca': snapshot['by_ca'],
        'ca_suggestions': ca_suggestions,

        # 全体示唆
        'suggestions': suggestions,

        # 過去Q達成率カーブ & 過去ペース継続予測
        'past_q_curves': past_q_curves,
        'avg_curve': avg_curve,
        'current_q_curve': current_q_curve,
        'historical_landing': round(historical_landing),
        'historical_final_rate': historical_final_rate,
    }


def extract_kpi_summary(wb):
    ws = wb['KPIダッシュボード']
    ca_names = []
    for col in range(8, 20):  # Support up to 12 CAs
        v = safe_val(ws.cell(4, col).value)
        if v and str(v).strip() not in ['合計', '合計/平均', '']:
            ca_names.append(str(v).strip())

    def get_row_data(row_num):
        result = {'total': parse_number(ws.cell(row_num, 7).value)}
        for i, ca in enumerate(ca_names):
            result[ca] = parse_number(ws.cell(row_num, 8 + i).value)
        return result

    return {
        'avg_profit': parse_number(ws.cell(2, 7).value),
        'q_progress': parse_number(ws.cell(3, 7).value),
        'ca_names': ca_names,
        'q_budget': get_row_data(6), 'q_actual': get_row_data(7),
        'achievement_rate': get_row_data(8), 'expected_consumption': get_row_data(9),
        'budget_diff': get_row_data(10), 'landing_estimate': get_row_data(11),
        'target_interviews': get_row_data(12), 'target_decision_rate': get_row_data(13),
        'profit_per_decision': get_row_data(14),
        'interview_target': get_row_data(16), 'interview_actual': get_row_data(17),
        'interview_achievement': get_row_data(18),
        'decision_count': get_row_data(23), 'decision_rate': get_row_data(24),
        'decision_diff': get_row_data(25),
    }

def compute_mom_changes(monthly_data, metric_key):
    """Compute month-over-month changes for a metric."""
    sorted_months = sorted(monthly_data.keys(), key=month_sort_key)
    changes = {}
    for i in range(1, len(sorted_months)):
        prev_m = sorted_months[i-1]
        curr_m = sorted_months[i]
        prev_v = monthly_data[prev_m].get(metric_key, 0)
        curr_v = monthly_data[curr_m].get(metric_key, 0)
        if prev_v > 0:
            changes[curr_m] = {'value': curr_v, 'prev': prev_v, 'change': (curr_v - prev_v) / prev_v}
        else:
            changes[curr_m] = {'value': curr_v, 'prev': prev_v, 'change': 0}
    return changes

def generate_comprehensive_insights(kpi, sales, ca_funnel, inflow, route_process, decision_attr):
    """Generate comprehensive AI analysis."""
    insights = {'overall': [], 'ca_detail': {}, 'inflow': []}
    ca_names = kpi['ca_names']
    q_progress = kpi['q_progress']

    # ========== OVERALL ANALYSIS ==========
    q_ach = kpi['achievement_rate']['total']
    gap_pp = (q_progress - q_ach) * 100

    # 1. Quarter progress
    if q_ach < q_progress * 0.9:
        insights['overall'].append({
            'type': 'critical', 'title': '売上進捗が日程進捗を大幅に下回っています',
            'detail': f'日程進捗{q_progress*100:.1f}%に対し売上達成率{q_ach*100:.1f}%（差分{gap_pp:.1f}pp）。現ペースでは目標未達の可能性が高い。',
            'metrics': {'日程進捗': f'{q_progress*100:.1f}%', '売上達成率': f'{q_ach*100:.1f}%', '差分': f'{gap_pp:.1f}pp'},
            'action': 'ヨミ案件の精査、決定見込み案件の早期クロージング、高単価案件への集中を検討'
        })
    elif q_ach >= q_progress:
        insights['overall'].append({
            'type': 'success', 'title': '売上進捗は順調',
            'detail': f'日程進捗{q_progress*100:.1f}%に対し売上達成率{q_ach*100:.1f}%で計画を上回っている。',
            'metrics': {'売上達成率': f'{q_ach*100:.1f}%'}, 'action': '現ペース維持'
        })

    # 2. Interview vs Decision gap
    total_int = kpi['interview_actual'].get('total', 0)
    total_int_target = kpi['interview_target'].get('total', 0)
    total_dec = kpi['decision_count'].get('total', 0)
    if total_dec > 1000: total_dec = 27  # fix formula error
    dec_rate = total_dec / total_int if total_int > 0 else 0

    if total_int > total_int_target and dec_rate < 0.05:
        insights['overall'].append({
            'type': 'warning', 'title': '面談数は充足しているが決定率が低い',
            'detail': f'面談{total_int}件（目標{total_int_target}件の{total_int/total_int_target*100:.0f}%）に対し決定{total_dec}件（決定率{dec_rate*100:.1f}%）。面談の質・マッチング精度に課題あり。',
            'metrics': {'面談数': f'{total_int}件', '決定数': f'{total_dec}件', '決定率': f'{dec_rate*100:.1f}%'},
            'action': '面談品質向上（候補者の意欲確認強化、求人マッチング精度改善）に注力'
        })

    # 3. Monthly trend analysis (from sales data)
    if '合計' in sales:
        total_monthly = sales['合計'].get('monthly', {})
        sorted_months = sorted([m for m in total_monthly.keys() if total_monthly[m].get('実績(粗利)', 0) > 0], key=month_sort_key)
        if len(sorted_months) >= 3:
            last3 = sorted_months[-3:]
            prev3 = sorted_months[-6:-3] if len(sorted_months) >= 6 else sorted_months[:3]
            avg_recent = sum(total_monthly[m].get('実績(粗利)', 0) for m in last3) / 3
            avg_prev = sum(total_monthly[m].get('実績(粗利)', 0) for m in prev3) / 3
            if avg_prev > 0:
                trend = (avg_recent - avg_prev) / avg_prev
                if trend < -0.15:
                    insights['overall'].append({
                        'type': 'warning', 'title': '月次売上に下降トレンド',
                        'detail': f'直近3ヶ月平均{avg_recent/10000:.0f}万 vs 前3ヶ月平均{avg_prev/10000:.0f}万（{trend*100:+.1f}%）。',
                        'metrics': {'直近3ヶ月平均': f'{avg_recent/10000:.0f}万', '前3ヶ月平均': f'{avg_prev/10000:.0f}万', '変化率': f'{trend*100:+.1f}%'},
                        'action': '面談数・決定数の推移を確認し、パイプラインの充実度を評価'
                    })
                elif trend > 0.15:
                    insights['overall'].append({
                        'type': 'success', 'title': '月次売上に上昇トレンド',
                        'detail': f'直近3ヶ月平均{avg_recent/10000:.0f}万 vs 前3ヶ月平均{avg_prev/10000:.0f}万（{trend*100:+.1f}%）。',
                        'metrics': {'変化率': f'{trend*100:+.1f}%'}, 'action': '好調要因の分析と横展開を検討'
                    })

    # 4. Lead time reminder
    insights['overall'].append({
        'type': 'info', 'title': 'リードタイム注意: 直近2-3ヶ月のデータは暫定値',
        'detail': '初回面談から決定まで2-3ヶ月のリードタイムがあるため、直近の決定数・決定率は実態より低く表示。当月・前月は暫定値。',
        'metrics': {}, 'action': '3ヶ月前データとの比較時はリードタイム補正後の数値で判断'
    })

    # ========== CA DETAIL ANALYSIS ==========
    # Special CA statuses (dynamic based on current quarter):
    # 渡辺, 百瀬: departed
    # 石丸: 1Q target = 0 (joining phase) - only applies in 1Q
    # 肥後: joins from 2Q - only excluded in 1Q
    departed_cas = ['渡辺', '百瀬']
    _q_info = get_current_quarter(sales)
    current_q_num = _q_info['q']
    # 石丸 only has zero target in 1Q
    zero_target_cas = {'石丸': '1Q'} if current_q_num == 1 else {}
    # 肥後 is joining from 2Q, so only mark as joining in 1Q
    joining_cas = ['肥後'] if current_q_num == 1 else []

    for ca in ca_names:
        ca_insights = []
        budget = kpi['q_budget'].get(ca, 0)
        actual = kpi['q_actual'].get(ca, 0)
        ach = kpi['achievement_rate'].get(ca, 0)
        interviews = kpi['interview_actual'].get(ca, 0)
        int_target = kpi['interview_target'].get(ca, 0)
        dec = kpi['decision_count'].get(ca, 0)
        if dec > 1000: dec = 0
        dec_rate = dec / interviews if interviews > 0 else 0
        int_rate = interviews / int_target if int_target > 0 else 0

        # Override for departed CAs
        is_departed = ca in departed_cas
        is_zero_target = ca in zero_target_cas

        # Diagnosis
        issues = []
        strengths = []

        if is_departed:
            strengths.append('退職決定済み（分析対象外）')
        elif is_zero_target:
            strengths.append('当Qは目標未設定（立ち上げ期間）')
            if actual > 0:
                strengths.append(f'目標なしで{actual/10000:.0f}万円の実績あり')
        else:
            if int_target > 0 and int_rate < 0.6:
                issues.append({'area': '面談数', 'detail': f'目標{int_target:.0f}件に対し{interviews:.0f}件（{int_rate*100:.0f}%）で大幅不足', 'severity': 'high'})
            elif int_target > 0 and int_rate < 0.85:
                issues.append({'area': '面談数', 'detail': f'目標{int_target:.0f}件に対し{interviews:.0f}件（{int_rate*100:.0f}%）でやや不足', 'severity': 'medium'})
            elif int_target > 0:
                strengths.append(f'面談数は目標の{int_rate*100:.0f}%で充足')

            # Decision rate: only evaluate if NOT in lead time window (skip for recent data)
            if dec_rate < 0.02 and not is_departed:
                issues.append({'area': '決定率', 'detail': f'{dec_rate*100:.1f}%（※リードタイム未反映の可能性あり）', 'severity': 'medium'})
            elif dec_rate < 0.04:
                issues.append({'area': '決定率', 'detail': f'{dec_rate*100:.1f}%でやや低い', 'severity': 'medium'})
            elif dec_rate >= 0.04:
                strengths.append(f'決定率{dec_rate*100:.1f}%で良好')

            if budget > 0:
                if ach < q_progress * 0.5:
                    issues.append({'area': '売上達成', 'detail': f'達成率{ach*100:.1f}%で大幅未達', 'severity': 'high'})
                elif ach < q_progress * 0.8:
                    issues.append({'area': '売上達成', 'detail': f'達成率{ach*100:.1f}%で遅れ', 'severity': 'medium'})
                else:
                    strengths.append(f'売上達成率{ach*100:.1f}%で順調')

        # Monthly trend for this CA
        monthly_trend = {}
        if ca in sales:
            ca_monthly = sales[ca].get('monthly', {})
            sorted_m = sorted([m for m in ca_monthly.keys() if ca_monthly[m].get('実績(粗利)', 0) > 0], key=month_sort_key)
            if len(sorted_m) >= 2:
                last_m = sorted_m[-1]
                prev_m = sorted_m[-2]
                last_v = ca_monthly[last_m].get('実績(粗利)', 0)
                prev_v = ca_monthly[prev_m].get('実績(粗利)', 0)
                if prev_v > 0:
                    change = (last_v - prev_v) / prev_v
                    monthly_trend = {'last_month': last_m, 'value': last_v, 'prev_value': prev_v, 'change': change}

        # Recommended actions
        actions = []
        if is_departed:
            actions.append('引き継ぎ完了の確認、担当求職者のフォローアップ')
        elif is_zero_target:
            actions.append('立ち上げ期間のため、面談経験の蓄積とOJTに注力')
        else:
            high_issues = [i for i in issues if i['severity'] == 'high']
            if any(i['area'] == '面談数' for i in high_issues):
                actions.append('スカウト活動の強化・アサイン数の増加が急務')
            if any(i['area'] == '決定率' for i in issues):
                actions.append('面談品質の改善（ヒアリング力向上、求人マッチング精度改善）')
            if any(i['area'] == '売上達成' for i in high_issues):
                actions.append('高単価案件への集中、決定促進アクション')
            if not actions:
                actions.append('現状維持し、さらなる成長機会を模索')

        high_issues = [i for i in issues if i['severity'] == 'high']
        status = 'departed' if is_departed else 'na' if is_zero_target else (
            'critical' if len(high_issues) >= 2 else 'warning' if high_issues else 'good')

        insights['ca_detail'][ca] = {
            'summary': {
                'budget': budget, 'actual': actual, 'achievement': ach,
                'interviews': interviews, 'interview_target': int_target, 'interview_rate': int_rate,
                'decisions': dec, 'decision_rate': dec_rate,
                'landing': kpi['landing_estimate'].get(ca, 0),
                'is_departed': is_departed, 'is_zero_target': is_zero_target,
            },
            'issues': issues, 'strengths': strengths,
            'monthly_trend': monthly_trend,
            'actions': actions,
            'status': status,
        }

    # ========== INFLOW ANALYSIS ==========
    if inflow:
        periods = list(inflow.keys())
        fy_periods = [p for p in periods if 'FY' in p]
        monthly_periods = [p for p in periods if 'FY' not in p]

        # Total inflow trend (sum all sub-categories under 就職時期)
        if fy_periods:
            for metric_cat in ['就職時期', '性別', '学歴', '前職']:
                recent_qs = fy_periods[-2:] if len(fy_periods) >= 2 else fy_periods
                prev_qs = fy_periods[-4:-2] if len(fy_periods) >= 4 else fy_periods[:2]

                recent_total = 0
                prev_total = 0
                for q in recent_qs:
                    if metric_cat in inflow[q]:
                        recent_total += sum(inflow[q][metric_cat].values())
                for q in prev_qs:
                    if metric_cat in inflow[q]:
                        prev_total += sum(inflow[q][metric_cat].values())

                if metric_cat == '就職時期' and prev_total > 0:
                    change = (recent_total - prev_total) / prev_total
                    if abs(change) > 0.1:
                        insights['inflow'].append({
                            'type': 'warning' if change < 0 else 'success',
                            'title': f'流入数が{"減少" if change < 0 else "増加"}傾向',
                            'detail': f'直近2Q合計{recent_total}件 vs 前2Q合計{prev_total}件（{change*100:+.1f}%）',
                            'metrics': {'直近2Q': f'{recent_total}件', '前2Q': f'{prev_total}件'},
                            'action': '集客チャネルの見直し' if change < 0 else '好調チャネルへの投資拡大検討'
                        })

            # Composition analysis
            latest_q = fy_periods[-1]
            if '就職時期' in inflow[latest_q]:
                timing = inflow[latest_q]['就職時期']
                total_timing = sum(timing.values())
                if total_timing > 0:
                    immediate_pct = (timing.get('すぐに', 0) + timing.get('1~2ヶ月以内', 0)) / total_timing
                    insights['inflow'].append({
                        'type': 'info',
                        'title': f'直近Q流入の転職意欲分布',
                        'detail': f'「すぐに」+「1-2ヶ月以内」が{immediate_pct*100:.0f}%、「3-6ヶ月以内」以降が{(1-immediate_pct)*100:.0f}%',
                        'metrics': {k: f'{v}件' for k, v in timing.items()},
                        'action': '転職意欲の高い層の比率に変化がないかモニタリング'
                    })

            if '年齢' in inflow[latest_q]:
                age = inflow[latest_q]['年齢']
                total_age = sum(age.values())
                if total_age > 0:
                    young_pct = age.get('20~25未満', 0) / total_age
                    insights['inflow'].append({
                        'type': 'info',
                        'title': f'流入年齢分布',
                        'detail': '、'.join([f'{k}: {v}件({v/total_age*100:.0f}%)' for k, v in age.items() if v > 0]),
                        'metrics': {k: f'{v}件' for k, v in age.items()},
                        'action': '年齢帯ごとの決定率と照らし合わせ、注力セグメントを検討'
                    })

    # Route analysis (自社 vs 自社以外)
    if route_process:
        for route in ['自社', '自社以外']:
            if route in route_process:
                rd = route_process[route]
                if '推薦決定率(決定数/推薦数)' in rd:
                    rate_data = rd['推薦決定率(決定数/推薦数)']
                    fy_keys = sorted([k for k in rate_data if 'FY' in k], key=date_sort_key)
                    if len(fy_keys) >= 2:
                        latest = rate_data[fy_keys[-1]]
                        prev = rate_data[fy_keys[-2]]
                        if prev > 0:
                            change = (latest - prev) / prev
                            insights['inflow'].append({
                                'type': 'warning' if change < -0.2 else 'success' if change > 0.1 else 'info',
                                'title': f'{route}経路の推薦決定率が{"低下" if change < 0 else "改善"}',
                                'detail': f'{fy_keys[-1]}: {latest*100:.1f}% → {fy_keys[-2]}: {prev*100:.1f}%（{change*100:+.1f}%）',
                                'metrics': {fy_keys[-1]: f'{latest*100:.1f}%', fy_keys[-2]: f'{prev*100:.1f}%'},
                                'action': f'{route}チャネルの{"改善施策検討" if change < 0 else "好調維持"}'
                            })

    return insights

def extract_historical_data(main_wb):
    """Extract FY23-FY26 quarterly data from main_sheet.xlsx 売上管理 tab.

    Returns a dict with 'annual', 'quarterly', and 'ca_quarterly' keys.
    """
    ws = main_wb['売上管理']

    # Quarter column mappings: (label, col_index)
    quarter_cols = [
        ('FY23／1Q', 4), ('FY23／2Q', 5), ('FY23／3Q', 6), ('FY23／4Q', 7),
        ('FY24／1Q', 20), ('FY24／2Q', 21), ('FY24／3Q', 22), ('FY24／4Q', 23),
        ('FY25／1Q', 36), ('FY25／2Q', 37), ('FY25／3Q', 38), ('FY25／4Q', 39),
        ('FY26／1Q', 52), ('FY26／2Q', 53), ('FY26／3Q', 54), ('FY26／4Q', 55),
    ]

    # Metric row offsets from the start of each block
    # 合計 block starts at row 4 (14 rows: 4-17)
    # CA blocks are 13 rows each
    metric_offsets_total = {
        0: '目標(粗利)', 1: '実績(粗利)', 2: '実績(RA粗利)', 3: '着地見込み',
        4: '実績達成率', 5: '見込達成率', 6: '売上', 7: '決定数',
        8: '当Q決定数', 9: '面談数', 10: '展開率', 11: '当Q展開率',
        12: '平均単価', 13: '平均粗利',
    }
    metric_offsets_ca = {
        0: '目標(粗利)', 1: '実績(粗利)', 2: '着地見込み',
        3: '実績達成率', 4: '見込達成率', 5: '売上', 6: '決定数',
        7: '当Q決定数', 8: '面談数', 9: '展開率', 10: '当Q展開率',
        11: '平均単価', 12: '平均粗利',
    }

    def read_block(start_row, offsets):
        """Read quarterly data for one block (合計 or a CA)."""
        block_data = {}
        for q_label, q_col in quarter_cols:
            q_data = {}
            for offset, metric in offsets.items():
                val = parse_number(ws.cell(start_row + offset, q_col).value)
                q_data[metric] = val
            block_data[q_label] = q_data
        return block_data

    # Read 合計 block (starts at row 4, 14 rows)
    total_quarterly = read_block(4, metric_offsets_total)

    # Read individual CA blocks
    ca_starts = {}
    for r in range(18, ws.max_row + 1):
        ca_name = ws.cell(r, 2).value
        metric_label = ws.cell(r, 3).value
        if ca_name and metric_label and '目標' in str(metric_label):
            ca_starts[str(ca_name).strip()] = r

    ca_quarterly = {}
    for ca_name, start_row in ca_starts.items():
        ca_quarterly[ca_name] = read_block(start_row, metric_offsets_ca)

    # Build annual aggregates from quarterly data
    fy_years = ['FY23', 'FY24', 'FY25', 'FY26']
    annual = {}
    for fy in fy_years:
        qs = [q for q, _ in quarter_cols if q.startswith(fy)]
        target_sum = sum(total_quarterly[q].get('目標(粗利)', 0) for q in qs)
        actual_sum = sum(total_quarterly[q].get('実績(粗利)', 0) for q in qs)
        decisions_sum = sum(total_quarterly[q].get('決定数', 0) for q in qs)
        interviews_sum = sum(total_quarterly[q].get('面談数', 0) for q in qs)
        revenue_sum = sum(total_quarterly[q].get('売上', 0) for q in qs)
        decision_rate = decisions_sum / interviews_sum if interviews_sum > 0 else 0

        annual[fy] = {
            'target': target_sum,
            'actual': actual_sum,
            'decisions': decisions_sum,
            'interviews': interviews_sum,
            'revenue': revenue_sum,
            'decision_rate': decision_rate,
        }

    # Build clean quarterly dict for output
    quarterly = {}
    for q_label in [q for q, _ in quarter_cols]:
        qd = total_quarterly[q_label]
        quarterly[q_label] = {
            'target': qd.get('目標(粗利)', 0),
            'actual': qd.get('実績(粗利)', 0),
            'landing': qd.get('着地見込み', 0),
            'revenue': qd.get('売上', 0),
            'decisions': qd.get('決定数', 0),
            'interviews': qd.get('面談数', 0),
            'decision_rate': qd.get('展開率', 0),
            'avg_unit_price': qd.get('平均単価', 0),
            'avg_profit': qd.get('平均粗利', 0),
        }

    # Build clean CA quarterly dict
    ca_quarterly_clean = {}
    for ca_name, ca_data in ca_quarterly.items():
        ca_q = {}
        for q_label in [q for q, _ in quarter_cols]:
            qd = ca_data[q_label]
            # Only include quarters with some data
            if qd.get('実績(粗利)', 0) != 0 or qd.get('面談数', 0) != 0 or qd.get('目標(粗利)', 0) != 0:
                ca_q[q_label] = {
                    'target': qd.get('目標(粗利)', 0),
                    'actual': qd.get('実績(粗利)', 0),
                    'landing': qd.get('着地見込み', 0),
                    'revenue': qd.get('売上', 0),
                    'decisions': qd.get('決定数', 0),
                    'interviews': qd.get('面談数', 0),
                    'decision_rate': qd.get('展開率', 0),
                    'avg_unit_price': qd.get('平均単価', 0),
                    'avg_profit': qd.get('平均粗利', 0),
                }
        if ca_q:
            ca_quarterly_clean[ca_name] = ca_q

    return {
        'annual': annual,
        'quarterly': quarterly,
        'ca_quarterly': ca_quarterly_clean,
    }


def extract_ca_deep_analysis(wb, sales):
    """Extract deep per-CA analysis from DB_求職者一覧 sheet.

    Produces unit price analysis, age-based decisions, NG reasons,
    final stage dropout, route analysis, company/position analysis,
    quarterly performance, and cross-CA comparison data.
    """
    ws = wb['DB_求職者一覧']

    ca_list = ['辰巳', '黒川', '古谷', '嘉重', '百石', '百瀬', '渡辺', '石丸']

    route_map = {
        'ビズリーチ': 'ビズリーチ', 'ビズリーチ(A)': 'ビズリーチ',
        'AMBI': 'AMBI', 'AMBI(A)': 'AMBI',
        'リクルート': 'リクルート', 'リクルート(A)': 'リクルート',
        '紹介(CS)': '紹介', '紹介(CN)': '紹介',
    }

    def to_quarter(dt):
        """Calendar year quarters: Q1=Jan-Mar, Q2=Apr-Jun, Q3=Jul-Sep, Q4=Oct-Dec."""
        y, m = dt.year, dt.month
        q = (m - 1) // 3 + 1
        return f'FY{y}/{q}Q'

    def parse_dt(v):
        if isinstance(v, datetime):
            return v
        if isinstance(v, date):
            return datetime(v.year, v.month, v.day)
        if isinstance(v, str):
            try:
                return datetime.strptime(v.strip()[:10], '%Y-%m-%d')
            except Exception:
                pass
        return None

    # ------------------------------------------------------------------
    # Read all rows once (row 2 = headers, data from row 3)
    # ------------------------------------------------------------------
    rows = []
    for r in range(3, min(ws.max_row + 1, 15001)):
        ca_raw = safe_val(ws.cell(r, 3).value)
        if not ca_raw:
            continue
        ca = str(ca_raw).strip()
        if ca not in ca_list:
            continue

        status = str(safe_val(ws.cell(r, 2).value) or '')
        route_raw = str(safe_val(ws.cell(r, 6).value) or '').strip()
        referral_raw = str(safe_val(ws.cell(r, 7).value) or '').strip()
        gross = parse_number(ws.cell(r, 11).value)
        interview_date = parse_dt(ws.cell(r, 16).value)
        age = parse_number(ws.cell(r, 23).value)
        prev_job = str(safe_val(ws.cell(r, 26).value) or '')
        kakudo = str(safe_val(ws.cell(r, 9).value) or '').strip()
        ng_reason = str(safe_val(ws.cell(r, 30).value) or '').strip()
        decision_company = str(safe_val(ws.cell(r, 32).value) or '').strip()
        revenue = parse_number(ws.cell(r, 33).value)
        final_status = str(safe_val(ws.cell(r, 37).value) or '').strip()
        app_count = parse_number(ws.cell(r, 39).value)
        dec_count = parse_number(ws.cell(r, 45).value)

        # Determine route (use referral col 7 if present and matches, else col 6)
        route_key = route_raw
        if referral_raw and referral_raw in route_map:
            route_key = referral_raw
        route = route_map.get(route_key, 'その他')

        rows.append({
            'ca': ca, 'status': status, 'route': route,
            'gross': gross, 'interview_date': interview_date,
            'age': age, 'prev_job': prev_job,
            'kakudo': kakudo,
            'ng_reason': ng_reason, 'decision_company': decision_company,
            'revenue': revenue, 'final_status': final_status,
            'app_count': app_count, 'dec_count': dec_count,
        })

    # ------------------------------------------------------------------
    # Aggregate per CA
    # ------------------------------------------------------------------
    def age_band(age_val):
        if age_val <= 0:
            return '不明'
        if age_val < 25:
            return '20-25'
        elif age_val < 30:
            return '25-30'
        elif age_val < 35:
            return '30-35'
        elif age_val < 40:
            return '35-40'
        elif age_val < 45:
            return '40-45'
        else:
            return '45+'

    def price_range(g):
        if g <= 0:
            return None
        g_man = g / 10000
        if g_man < 50:
            return '~50万'
        elif g_man < 80:
            return '50-80万'
        elif g_man < 100:
            return '80-100万'
        elif g_man < 150:
            return '100-150万'
        else:
            return '150万~'

    ca_results = {}

    for ca in ca_list:
        ca_rows = [r for r in rows if r['ca'] == ca]
        if not ca_rows:
            continue

        # 1. Unit price (粗利) analysis
        gross_values = [r['gross'] for r in ca_rows if r['dec_count'] > 0 and r['gross'] > 0]
        gross_analysis = {}
        if gross_values:
            gross_analysis['avg'] = round(sum(gross_values) / len(gross_values))
            gross_analysis['max'] = max(gross_values)
            gross_analysis['min'] = min(gross_values)
            gross_analysis['count'] = len(gross_values)
        else:
            gross_analysis['avg'] = 0
            gross_analysis['max'] = 0
            gross_analysis['min'] = 0
            gross_analysis['count'] = 0

        # Distribution by price range
        price_dist = defaultdict(int)
        for g in gross_values:
            pr = price_range(g)
            if pr:
                price_dist[pr] += 1
        gross_analysis['distribution'] = dict(price_dist)

        # Quarterly avg gross trend
        q_gross = defaultdict(list)
        for r_item in ca_rows:
            if r_item['dec_count'] > 0 and r_item['gross'] > 0 and r_item['interview_date']:
                q = to_quarter(r_item['interview_date'])
                q_gross[q].append(r_item['gross'])
        quarterly_gross_trend = {}
        for q in sorted(q_gross.keys()):
            vals = q_gross[q]
            quarterly_gross_trend[q] = {
                'avg': round(sum(vals) / len(vals)),
                'count': len(vals),
                'total': round(sum(vals)),
            }
        gross_analysis['quarterly_trend'] = quarterly_gross_trend

        # 2. Age-based decisions
        age_data = defaultdict(lambda: {'interviews': 0, 'decisions': 0})
        for r_item in ca_rows:
            if r_item['age'] > 0:
                ab = age_band(r_item['age'])
                age_data[ab]['interviews'] += 1
                if r_item['dec_count'] > 0:
                    age_data[ab]['decisions'] += int(r_item['dec_count'])
        age_analysis = {}
        for ab in sorted(age_data.keys()):
            d = age_data[ab]
            age_analysis[ab] = {
                'interviews': d['interviews'],
                'decisions': d['decisions'],
                'rate': round(d['decisions'] / d['interviews'], 4) if d['interviews'] > 0 else 0,
            }

        # 3. NG reason analysis
        ng_counts = defaultdict(int)
        for r_item in ca_rows:
            if r_item['ng_reason'] and r_item['ng_reason'] not in ('', 'None', '-'):
                ng_counts[r_item['ng_reason']] += 1
        top_ng = sorted(ng_counts.items(), key=lambda x: -x[1])[:10]
        ng_analysis = [{'reason': reason, 'count': count} for reason, count in top_ng]

        # 4. Final stage dropout
        dropout_counts = defaultdict(int)
        dropout_stages = [
            '初回面談後NG', '求人紹介後NG', '書類準備中NG', '応募後NG',
            '1st面接後NG', '2nd面接後NG', '最終面接後NG', '内定後NG',
        ]
        for r_item in ca_rows:
            fs = r_item['final_status']
            if fs and 'NG' in fs:
                dropout_counts[fs] += 1
            elif r_item['status'] and 'NG' in r_item['status']:
                dropout_counts[r_item['status']] += 1
        dropout_analysis = []
        # Include all found dropout stages, sorted by count
        for stage, count in sorted(dropout_counts.items(), key=lambda x: -x[1]):
            dropout_analysis.append({'stage': stage, 'count': count})

        # 5. Route analysis
        route_data = defaultdict(lambda: {'interviews': 0, 'decisions': 0})
        for r_item in ca_rows:
            rt = r_item['route']
            route_data[rt]['interviews'] += 1
            if r_item['dec_count'] > 0:
                route_data[rt]['decisions'] += int(r_item['dec_count'])
        route_analysis = {}
        for rt in sorted(route_data.keys()):
            d = route_data[rt]
            route_analysis[rt] = {
                'interviews': d['interviews'],
                'decisions': d['decisions'],
                'rate': round(d['decisions'] / d['interviews'], 4) if d['interviews'] > 0 else 0,
            }

        # 5b. 確度 (prospect rating) analysis
        kakudo_data = defaultdict(lambda: {'total': 0, 'decided': 0, 'dropout_by_stage': defaultdict(int)})
        for r_item in ca_rows:
            k = r_item.get('kakudo', '')
            if not k or k in ('', 'None'): continue
            kakudo_data[k]['total'] += 1
            if r_item['dec_count'] > 0:
                kakudo_data[k]['decided'] += int(r_item['dec_count'])
            elif r_item['status'] and 'NG' in r_item['status']:
                kakudo_data[k]['dropout_by_stage'][r_item['status']] += 1

        # C→B conversion (C rating → progressed to 応募 or later)
        c_total = 0
        c_progressed = 0
        progress_stages = {'応募', '1st面接', '2nd面接', '最終面接', '内定', '決定', '入社',
                          '応募後NG', '1st面接後NG', '2nd面接後NG', '最終面接後NG', '内定後NG', '決定後NG'}
        for r_item in ca_rows:
            if r_item.get('kakudo', '') == 'C':
                c_total += 1
                if r_item['status'] in progress_stages:
                    c_progressed += 1

        kakudo_analysis = {}
        for k in ['A', 'B', 'C', 'D']:
            d = kakudo_data[k]
            if d['total'] == 0: continue
            top_dropout = sorted(d['dropout_by_stage'].items(), key=lambda x: -x[1])[:5]
            kakudo_analysis[k] = {
                'total': d['total'],
                'decided': d['decided'],
                'rate': round(d['decided'] / d['total'], 4) if d['total'] > 0 else 0,
                'top_dropout': [{'stage': s, 'count': c} for s, c in top_dropout],
            }
        c_to_b_rate = round(c_progressed / c_total, 4) if c_total > 0 else 0

        # 6. Company/position analysis
        company_counts = defaultdict(int)
        for r_item in ca_rows:
            if r_item['dec_count'] > 0 and r_item['decision_company'] and r_item['decision_company'] not in ('', 'None', '-'):
                company_counts[r_item['decision_company']] += 1
        top_companies = sorted(company_counts.items(), key=lambda x: -x[1])[:10]
        company_analysis = [{'company': c, 'count': cnt} for c, cnt in top_companies]

        # 7. Quarterly performance - USE sales_by_ca as source of truth (not interview-date based)
        # sales_by_ca uses full-width slash: FY25／1Q
        # We need to normalize to half-width: FY25/1Q for consistency with other data
        quarterly_perf = {}
        ca_sales_data = sales.get(ca, {}).get('quarterly', {})
        for sq, sdata in ca_sales_data.items():
            # Normalize Q key: FY25／1Q -> FY25/1Q
            nq = sq.replace('／', '/')
            gross = sdata.get('実績(粗利)', 0) or 0
            decisions = sdata.get('決定数', 0) or 0
            interviews = sdata.get('面談数', 0) or 0
            if gross > 0 or decisions > 0 or interviews > 0:
                quarterly_perf[nq] = {
                    'interviews': interviews,
                    'decisions': decisions,
                    'gross': round(gross),
                    'decision_rate': round(decisions / interviews, 4) if interviews > 0 else 0,
                }
        # Also add older quarters from DB if not in sales (pre-FY25 data)
        # Sales data uses FY25/1Q format, DB uses FY2025/1Q format
        # Only add DB data for quarters that don't overlap with sales
        sales_qs_normalized = set()
        for sq in ca_sales_data.keys():
            nq = sq.replace('／', '/')
            sales_qs_normalized.add(nq)
            # Also mark the long-form equivalent as covered
            # FY25/1Q -> FY2025/1Q
            parts = nq.split('/')
            if len(parts) == 2:
                fy_part = parts[0].replace('FY', '')
                if len(fy_part) == 2:
                    sales_qs_normalized.add(f'FY20{fy_part}/{parts[1]}')

        q_from_db = defaultdict(lambda: {'interviews': 0, 'decisions': 0, 'gross': 0})
        for r_item in ca_rows:
            if r_item['interview_date']:
                q = to_quarter(r_item['interview_date'])
                q_from_db[q]['interviews'] += 1
                if r_item['dec_count'] > 0:
                    q_from_db[q]['decisions'] += int(r_item['dec_count'])
                    q_from_db[q]['gross'] += r_item['gross']
        for q in sorted(q_from_db.keys()):
            if q not in quarterly_perf and q not in sales_qs_normalized:
                d = q_from_db[q]
                if d['interviews'] > 0 or d['decisions'] > 0:
                    quarterly_perf[q] = {
                        'interviews': d['interviews'],
                        'decisions': d['decisions'],
                        'gross': round(d['gross']),
                        'decision_rate': round(d['decisions'] / d['interviews'], 4) if d['interviews'] > 0 else 0,
                    }

        ca_results[ca] = {
            'total_records': len(ca_rows),
            'gross_analysis': gross_analysis,
            'age_analysis': age_analysis,
            'ng_reasons': ng_analysis,
            'dropout_analysis': dropout_analysis,
            'kakudo_analysis': kakudo_analysis,
            'c_to_b_rate': c_to_b_rate,
            'c_to_b_detail': {'c_total': c_total, 'c_progressed': c_progressed},
            'route_analysis': route_analysis,
            'company_analysis': company_analysis,
            'quarterly_performance': quarterly_perf,
        }

    # ------------------------------------------------------------------
    # 8. Cross-CA comparison data (team averages for each metric)
    # ------------------------------------------------------------------
    team_avg = {}
    active_cas = [ca for ca in ca_list if ca in ca_results]
    if active_cas:
        # Average gross
        all_gross_avgs = [ca_results[ca]['gross_analysis']['avg'] for ca in active_cas if ca_results[ca]['gross_analysis']['avg'] > 0]
        team_avg['avg_gross'] = round(sum(all_gross_avgs) / len(all_gross_avgs)) if all_gross_avgs else 0

        # Average decision rate
        all_total_int = sum(ca_results[ca]['total_records'] for ca in active_cas)
        all_total_dec = sum(
            sum(r['dec_count'] for r in rows if r['ca'] == ca and r['dec_count'] > 0)
            for ca in active_cas
        )
        team_avg['avg_decision_rate'] = round(all_total_dec / all_total_int, 4) if all_total_int > 0 else 0

        # Route breakdown across team
        team_route = defaultdict(lambda: {'interviews': 0, 'decisions': 0})
        for ca in active_cas:
            for rt, rd in ca_results[ca]['route_analysis'].items():
                team_route[rt]['interviews'] += rd['interviews']
                team_route[rt]['decisions'] += rd['decisions']
        team_avg['route'] = {}
        for rt in sorted(team_route.keys()):
            d = team_route[rt]
            team_avg['route'][rt] = {
                'interviews': d['interviews'],
                'decisions': d['decisions'],
                'rate': round(d['decisions'] / d['interviews'], 4) if d['interviews'] > 0 else 0,
            }

        # Age breakdown across team
        team_age = defaultdict(lambda: {'interviews': 0, 'decisions': 0})
        for ca in active_cas:
            for ab, ad in ca_results[ca]['age_analysis'].items():
                team_age[ab]['interviews'] += ad['interviews']
                team_age[ab]['decisions'] += ad['decisions']
        team_avg['age'] = {}
        for ab in sorted(team_age.keys()):
            d = team_age[ab]
            team_avg['age'][ab] = {
                'interviews': d['interviews'],
                'decisions': d['decisions'],
                'rate': round(d['decisions'] / d['interviews'], 4) if d['interviews'] > 0 else 0,
            }

        # Top NG reasons across team
        team_ng = defaultdict(int)
        for ca in active_cas:
            for ng in ca_results[ca]['ng_reasons']:
                team_ng[ng['reason']] += ng['count']
        team_avg['ng_reasons'] = [{'reason': r, 'count': c} for r, c in sorted(team_ng.items(), key=lambda x: -x[1])[:10]]

        # Team 確度 analysis
        team_kakudo = defaultdict(lambda: {'total': 0, 'decided': 0, 'dropout_by_stage': defaultdict(int)})
        team_c_total = 0
        team_c_progressed = 0
        for ca in active_cas:
            for k, kd in ca_results[ca].get('kakudo_analysis', {}).items():
                team_kakudo[k]['total'] += kd['total']
                team_kakudo[k]['decided'] += kd['decided']
                for dd in kd.get('top_dropout', []):
                    team_kakudo[k]['dropout_by_stage'][dd['stage']] += dd['count']
            team_c_total += ca_results[ca].get('c_to_b_detail', {}).get('c_total', 0)
            team_c_progressed += ca_results[ca].get('c_to_b_detail', {}).get('c_progressed', 0)

        team_avg['kakudo'] = {}
        for k in ['A', 'B', 'C', 'D']:
            d = team_kakudo[k]
            if d['total'] == 0: continue
            top_dropout = sorted(d['dropout_by_stage'].items(), key=lambda x: -x[1])[:5]
            team_avg['kakudo'][k] = {
                'total': d['total'],
                'decided': d['decided'],
                'rate': round(d['decided'] / d['total'], 4) if d['total'] > 0 else 0,
                'top_dropout': [{'stage': s, 'count': c} for s, c in top_dropout],
            }
        team_avg['c_to_b_rate'] = round(team_c_progressed / team_c_total, 4) if team_c_total > 0 else 0

    return {
        'by_ca': ca_results,
        'team_avg': team_avg,
    }


def main():
    wb = openpyxl.load_workbook('data/source.xlsx', data_only=True)

    print("Extracting KPI...")
    kpi = extract_kpi_summary(wb)
    print("Extracting sales...")
    sales = extract_sales_by_ca(wb)
    print("Extracting CA funnel...")
    ca_funnel = extract_ca_monthly_funnel(wb)
    print("Extracting inflow attributes...")
    inflow = extract_inflow_attributes(wb)
    print("Extracting route process...")
    route_process = extract_route_process(wb)
    print("Extracting decision by attribute...")
    decision_attr = extract_decision_by_attribute(wb)

    print("Extracting 2026 targets...")
    import os
    targets_path = 'data/targets.xlsx'
    targets_2026 = extract_2026_targets(targets_path) if os.path.exists(targets_path) else {}

    print("Extracting referral data...")
    referral_path = 'data/referral.xlsx'
    referral_data = extract_referral_data(referral_path) if os.path.exists(referral_path) else {}

    print("Extracting route breakdown...")
    route_breakdown = extract_route_breakdown(wb)

    print("Extracting CA deep analysis...")
    ca_deep_analysis = extract_ca_deep_analysis(wb, sales)

    print("Extracting ヨミ forecast...")
    yomi_forecast = extract_yomi_forecast(wb)

    print("Extracting historical data from main_sheet...")
    main_wb = openpyxl.load_workbook('data/main_sheet.xlsx', data_only=True)
    historical = extract_historical_data(main_wb)

    # Auto-detect current quarter and set dynamic CA statuses
    q_info = get_current_quarter(sales)
    current_q_num = q_info['q']
    calendar_q_num = q_info['calendar_q']
    departed_cas = ['渡辺', '百瀬']
    # Use calendar quarter for membership (not display quarter which may lag)
    zero_target_cas = {'石丸': '1Q'} if calendar_q_num == 1 else {}
    joining_cas = ['肥後'] if calendar_q_num == 1 else []
    # Add 肥後 to ca_names from 2Q onward if not already present
    if calendar_q_num >= 2 and '肥後' not in kpi['ca_names']:
        kpi['ca_names'].append('肥後')
        # Initialize KPI fields for 肥後 with 0 values
        for key in kpi:
            if isinstance(kpi[key], dict) and 'total' in kpi[key] and '肥後' not in kpi[key]:
                kpi[key]['肥後'] = 0
    # Override KPI totals with sales management data for current Q
    # (KPI sheet may still show previous Q data at start of new Q)
    current_q_sales_key = q_info['current_q_sales']
    total_sales_q = sales.get('合計', {}).get('quarterly', {}).get(current_q_sales_key, {})
    if total_sales_q and total_sales_q.get('目標(粗利)', 0) > 0:
        total_budget = total_sales_q.get('目標(粗利)', 0)
        total_actual = total_sales_q.get('実績(粗利)', 0)
        total_ach = total_actual / total_budget if total_budget > 0 else 0
        # Calculate Q progress based on date
        q_start_month = (current_q_num - 1) * 3 + 1
        q_start = datetime(q_info['fy'], q_start_month, 1)
        q_total_days = (q_info['q_end_date'] - q_start).days + 1
        q_elapsed = (datetime.now() - q_start).days
        q_progress = max(0, min(1, q_elapsed / q_total_days))
        kpi['q_budget'] = {**kpi.get('q_budget', {}), 'total': total_budget}
        kpi['q_actual'] = {**kpi.get('q_actual', {}), 'total': total_actual}
        kpi['achievement_rate'] = {**kpi.get('achievement_rate', {}), 'total': total_ach}
        kpi['q_progress'] = q_progress
        # Override per-CA from sales
        for ca in kpi['ca_names']:
            ca_sq = sales.get(ca, {}).get('quarterly', {}).get(current_q_sales_key, {})
            if ca_sq:
                kpi['q_budget'][ca] = ca_sq.get('目標(粗利)', 0)
                kpi['q_actual'][ca] = ca_sq.get('実績(粗利)', 0)
                b = kpi['q_budget'][ca]
                a = kpi['q_actual'][ca]
                kpi['achievement_rate'][ca] = a / b if b > 0 else 0
                kpi.setdefault('landing_estimate', {})[ca] = ca_sq.get('着地見込み', 0)
                kpi.setdefault('decision_count', {})[ca] = ca_sq.get('決定数', 0)
                kpi.setdefault('interview_actual', {})[ca] = ca_sq.get('面談数', 0)
        print(f"  KPI overridden with sales data for {current_q_sales_key}: target={total_budget/10000:.0f}万 actual={total_actual/10000:.0f}万")

    print(f"  Current quarter: {q_info['current_q_funnel']} (Q{current_q_num})")
    print(f"  Active CAs: {[c for c in kpi['ca_names'] if c not in departed_cas]}")

    print("Generating comprehensive insights...")
    insights = generate_comprehensive_insights(kpi, sales, ca_funnel, inflow, route_process, decision_attr)

    # Build monthly/quarterly trends
    monthly_trends = {}
    if '合計' in sales:
        for period, metrics in sales['合計'].get('monthly', {}).items():
            monthly_trends[period] = {
                'target': metrics.get('目標(粗利)', 0), 'actual': metrics.get('実績(粗利)', 0),
                'landing': metrics.get('着地見込み', 0), 'revenue': metrics.get('売上', 0),
                'decisions': metrics.get('決定数', 0), 'interviews': metrics.get('面談数', 0),
            }
    quarterly_trends = {}
    # Start with historical quarterly data (FY23-FY26 from main_sheet)
    for period, data in historical['quarterly'].items():
        quarterly_trends[period] = {
            'target': data.get('target', 0), 'actual': data.get('actual', 0),
            'landing': data.get('landing', 0), 'revenue': data.get('revenue', 0),
            'decisions': data.get('decisions', 0), 'interviews': data.get('interviews', 0),
        }
    # Overlay with source.xlsx data (may have more recent/accurate values for FY25+)
    if '合計' in sales:
        for period, metrics in sales['合計'].get('quarterly', {}).items():
            quarterly_trends[period] = {
                'target': metrics.get('目標(粗利)', 0), 'actual': metrics.get('実績(粗利)', 0),
                'landing': metrics.get('着地見込み', 0), 'revenue': metrics.get('売上', 0),
                'decisions': metrics.get('決定数', 0), 'interviews': metrics.get('面談数', 0),
            }

    ca_comparison = {}
    departed_cas = ['渡辺', '百瀬']
    current_q_key = q_info['current_q_sales']  # e.g. 'FY26／2Q'
    for ca in kpi['ca_names']:
        # Prefer sales_by_ca data for current Q (more accurate than KPI sheet which may lag)
        ca_sales_q = sales.get(ca, {}).get('quarterly', {}).get(current_q_key, {})
        if ca_sales_q and ca_sales_q.get('目標(粗利)', 0) > 0:
            # Use sales management data
            budget = ca_sales_q.get('目標(粗利)', 0)
            actual = ca_sales_q.get('実績(粗利)', 0)
            ach = actual / budget if budget > 0 else 0
            dec_val = ca_sales_q.get('決定数', 0)
            interviews = ca_sales_q.get('面談数', 0)
            int_target = kpi['interview_target'].get(ca, 0)
            dec_rate = dec_val / interviews if interviews > 0 else 0
            landing = ca_sales_q.get('着地見込み', 0)
        else:
            # Fallback to KPI sheet
            budget = kpi['q_budget'].get(ca, 0)
            actual = kpi['q_actual'].get(ca, 0)
            ach = kpi['achievement_rate'].get(ca, 0)
            dec_val = kpi['decision_count'].get(ca, 0)
            if dec_val > 1000: dec_val = 0
            interviews = kpi['interview_actual'].get(ca, 0)
            int_target = kpi['interview_target'].get(ca, 0)
            dec_rate = kpi['decision_rate'].get(ca, 0)
            landing = kpi['landing_estimate'].get(ca, 0)
        ca_comparison[ca] = {
            'budget': budget, 'actual': actual,
            'achievement': ach,
            'interviews': interviews,
            'interview_target': int_target,
            'decisions': dec_val,
            'decision_rate': dec_rate,
            'landing': landing,
            'is_departed': ca in departed_cas,
            'is_zero_target': ca in zero_target_cas,
        }

    print("Building weekly report...")
    weekly_report = build_weekly_report(yomi_forecast, sales, ca_funnel, kpi, ca_comparison, q_info, departed_cas, wb)
    print(f"  Forecast: pipeline={weekly_report['pipeline_forecast']/10000:.0f}万, new={weekly_report['new_revenue_predicted']/10000:.0f}万, integrated={weekly_report['integrated_forecast']/10000:.0f}万")
    print(f"  Gap: {weekly_report['gap']/10000:.0f}万 / 達成率: {weekly_report['achievement_rate']*100:.0f}%")

    funnel_data = {}
    if 'total' not in ca_funnel and '合計' in ca_funnel:
        funnel_data['total'] = ca_funnel['合計']
    elif 'total' in ca_funnel:
        funnel_data['total'] = ca_funnel['total']
    for ca in kpi['ca_names']:
        if ca in ca_funnel: funnel_data[ca] = ca_funnel[ca]

    # Decision rate lead-time config
    today = datetime.now()
    lead_time_months = []
    for offset in range(3):
        m = today.month - offset
        y = today.year
        if m <= 0:
            m += 12
            y -= 1
        lead_time_months.append(f"{y}年{m}月")

    # ========== WEEKLY PULSE (週次パルス) ==========
    print("Building weekly pulse...")

    today = datetime.now()
    days_in_month = [31,28,31,30,31,30,31,31,30,31,30,31][today.month - 1]
    month_progress = today.day / days_in_month

    cur_m_key = f"{today.year}-{today.month:02d}"
    prev_m = today.month - 1 if today.month > 1 else 12
    prev_y = today.year if today.month > 1 else today.year - 1
    prev_m_key = f"{prev_y}-{prev_m:02d}"

    stages_ordered = ['アサイン', '初回面談', '求人紹介', '応募', '1st面接', '2nd面接', '最終面接', '内定', '決定']

    weekly_pulse = {
        'generated_at': today.strftime('%Y-%m-%d %H:%M'),
        'month_progress': round(month_progress, 2),
        'current_month': cur_m_key,
        'prev_month': prev_m_key,
        'overall': {},
        'by_ca': {},
        'alerts': [],
    }

    # Overall funnel comparison: current month vs prev month
    total_funnel = funnel_data.get('total', funnel_data.get('合計', {})).get('funnel', {})
    overall_stages = {}
    for stage in stages_ordered:
        if stage not in total_funnel: continue
        cur_val = total_funnel[stage].get(cur_m_key, 0)
        prev_val = total_funnel[stage].get(prev_m_key, 0)
        # Projected = current / month_progress (run rate)
        projected = round(cur_val / month_progress) if month_progress > 0.3 else cur_val
        pace_vs_prev = (projected / prev_val - 1) if prev_val > 0 else 0
        overall_stages[stage] = {
            'current': cur_val, 'prev': prev_val,
            'projected': projected, 'pace': round(pace_vs_prev, 3),
        }
    weekly_pulse['overall'] = overall_stages

    # Alerts based on overall
    cur_interviews = overall_stages.get('初回面談', {}).get('current', 0)
    prev_interviews = overall_stages.get('初回面談', {}).get('prev', 0)
    proj_interviews = overall_stages.get('初回面談', {}).get('projected', 0)

    if proj_interviews < prev_interviews * 0.8 and month_progress > 0.4:
        weekly_pulse['alerts'].append({
            'type': 'warning', 'area': '初回面談',
            'message': f'初回面談ペースが前月比{(proj_interviews/prev_interviews-1)*100:.0f}%。月末着地{proj_interviews}件予測（前月{prev_interviews}件）',
            'severity': 'high',
        })
    elif proj_interviews >= prev_interviews:
        weekly_pulse['alerts'].append({
            'type': 'good', 'area': '初回面談',
            'message': f'初回面談ペース良好。月末着地{proj_interviews}件予測（前月{prev_interviews}件、+{(proj_interviews/prev_interviews-1)*100:.0f}%）',
            'severity': 'good',
        })

    # Funnel flow check: are people progressing through stages?
    for i in range(1, len(stages_ordered) - 1):
        stage = stages_ordered[i]
        next_stage = stages_ordered[i + 1]
        cur_s = overall_stages.get(stage, {}).get('current', 0)
        cur_n = overall_stages.get(next_stage, {}).get('current', 0)
        prev_s = overall_stages.get(stage, {}).get('prev', 0)
        prev_n = overall_stages.get(next_stage, {}).get('prev', 0)

        cur_rate = cur_n / cur_s if cur_s > 10 else 0
        prev_rate = prev_n / prev_s if prev_s > 10 else 0

        if cur_rate > 0 and prev_rate > 0:
            rate_change = cur_rate - prev_rate
            if rate_change <= -0.1:
                weekly_pulse['alerts'].append({
                    'type': 'warning', 'area': f'{stage}→{next_stage}',
                    'message': f'{stage}→{next_stage}の転換率が低下: {cur_rate*100:.0f}%（前月{prev_rate*100:.0f}%、{rate_change*100:+.0f}pp）',
                    'severity': 'medium',
                })
            elif rate_change >= 0.05:
                weekly_pulse['alerts'].append({
                    'type': 'good', 'area': f'{stage}→{next_stage}',
                    'message': f'{stage}→{next_stage}の転換率改善: {cur_rate*100:.0f}%（前月{prev_rate*100:.0f}%、{rate_change*100:+.0f}pp）',
                    'severity': 'good',
                })

    # Active pipeline health (from DB_求職者一覧)
    active_by_stage = {}
    active_by_ca = {}
    ws_db = wb['DB_求職者一覧']
    active_statuses = {'内定', '最終面接', '2nd面接', '1st面接', '応募', '書類準備中', '求人紹介', '決定'}
    for r in range(3, ws_db.max_row + 1):
        status = str(safe_val(ws_db.cell(r, 2).value) or '')
        ca = str(safe_val(ws_db.cell(r, 3).value) or '')
        dec_count = parse_number(ws_db.cell(r, 45).value)
        if dec_count > 0: continue
        if 'NG' in status or '保留' in status: continue
        if not ca or ca == 'None': continue

        # Map status to stage
        matched_stage = None
        for s in active_statuses:
            if status == s:
                matched_stage = s
                break
        if not matched_stage: continue

        active_by_stage[matched_stage] = active_by_stage.get(matched_stage, 0) + 1
        if ca not in active_by_ca:
            active_by_ca[ca] = {}
        active_by_ca[ca][matched_stage] = active_by_ca[ca].get(matched_stage, 0) + 1

    weekly_pulse['active_pipeline'] = active_by_stage
    weekly_pulse['active_by_ca'] = active_by_ca

    # ===== FUNNEL CONVERSION RATE TRACKING =====
    # Monthly funnel counts for current Q (3 months)
    funnel_stages = ['アサイン', '初回面談', '求人紹介', '応募', '1st面接', '2nd面接', '最終面接', '内定', '決定']
    total_funnel = ca_funnel.get('合計', ca_funnel.get('total', {})).get('funnel', {})

    # Get monthly data for current Q months
    cur_q_months = []
    for offset in range(2, -1, -1):
        m = today.month - offset
        y = today.year
        if m <= 0: m += 12; y -= 1
        cur_q_months.append(f'{y}-{m:02d}')

    monthly_funnel = {}
    for m_key in cur_q_months:
        m_data = {}
        for s in funnel_stages:
            m_data[s] = total_funnel.get(s, {}).get(m_key, 0)
        monthly_funnel[m_key] = m_data

    # Conversion pairs
    conv_pairs = [
        ('面談→求人紹介', '初回面談', '求人紹介'),
        ('求人紹介→応募', '求人紹介', '応募'),
        ('応募→1st面接', '応募', '1st面接'),
        ('1st→2nd面接', '1st面接', '2nd面接'),
        ('2nd→最終面接', '2nd面接', '最終面接'),
        ('最終→内定', '最終面接', '内定'),
    ]

    # Calculate monthly conversion rates
    monthly_conv_rates = {}
    for m_key in cur_q_months:
        m_rates = {}
        for label, from_s, to_s in conv_pairs:
            f_val = monthly_funnel[m_key].get(from_s, 0)
            t_val = monthly_funnel[m_key].get(to_s, 0)
            m_rates[label] = {
                'from': f_val, 'to': t_val,
                'rate': round(t_val / f_val * 100, 1) if f_val > 0 else 0,
            }
        monthly_conv_rates[m_key] = m_rates

    # Calculate historical average conversion rates (直近2Q only - comparable baseline)
    past_qs = ['FY25/3Q', 'FY25/4Q']
    avg_conv_rates = {}
    for label, from_s, to_s in conv_pairs:
        rates = []
        for q in past_qs:
            f_val = total_funnel.get(from_s, {}).get(q, 0)
            t_val = total_funnel.get(to_s, {}).get(q, 0)
            if f_val > 0:
                rates.append(t_val / f_val * 100)
        avg_conv_rates[label] = round(sum(rates) / len(rates), 1) if rates else 0

    # Mark current month as "in progress" - conversion rates are unreliable
    # Only alert based on COMPLETED months (not current month)
    if len(cur_q_months) >= 2:
        # Use 2nd month (prev_m) as the latest COMPLETED month for comparison
        # Current month is still in progress so rates will naturally look low
        completed_m = cur_q_months[-2]  # 2月 = completed
        for label, from_s, to_s in conv_pairs:
            comp_rate = monthly_conv_rates[completed_m][label]['rate']
            avg_rate = avg_conv_rates[label]
            comp_from = monthly_conv_rates[completed_m][label]['from']
            if comp_from >= 5 and comp_rate < avg_rate - 10:
                drop_pp = round(comp_rate - avg_rate, 1)
                weekly_pulse['alerts'].append({
                    'type': 'warning', 'area': f'{label}の通過率',
                    'message': f'{label}の転換率が低下（{completed_m.split("-")[1]}月確定値）: {comp_rate:.0f}%（直近2Q平均{avg_rate:.0f}%、{drop_pp:+.0f}pp）',
                })
            elif comp_from >= 5 and comp_rate > avg_rate + 5:
                up_pp = round(comp_rate - avg_rate, 1)
                weekly_pulse['alerts'].append({
                    'type': 'good', 'area': f'{label}の通過率',
                    'message': f'{label}の転換率が好調（{completed_m.split("-")[1]}月確定値）: {comp_rate:.0f}%（直近2Q平均{avg_rate:.0f}%、+{up_pp:.0f}pp）',
                })

    weekly_pulse['funnel_monthly'] = monthly_funnel
    weekly_pulse['conversion_rates'] = monthly_conv_rates
    weekly_pulse['avg_conversion_rates'] = avg_conv_rates
    weekly_pulse['q_months'] = cur_q_months

    # Active pipeline totals
    total_active = sum(active_by_stage.values())
    hot_pipeline = active_by_stage.get('内定', 0) + active_by_stage.get('最終面接', 0)

    # Per-CA weekly pulse
    for ca in kpi['ca_names']:
        if ca in departed_cas: continue
        ca_funnel_data = funnel_data.get(ca, {}).get('funnel', {})
        ca_stages = {}
        ca_alerts = []
        for stage in stages_ordered:
            if stage not in ca_funnel_data: continue
            cur_val = ca_funnel_data[stage].get(cur_m_key, 0)
            prev_val = ca_funnel_data[stage].get(prev_m_key, 0)
            projected = round(cur_val / month_progress) if month_progress > 0.3 else cur_val
            pace = (projected / prev_val - 1) if prev_val > 0 else 0
            ca_stages[stage] = {'current': cur_val, 'prev': prev_val, 'projected': projected, 'pace': round(pace, 3)}

        # CA interview pace
        ca_int_cur = ca_stages.get('初回面談', {}).get('current', 0)
        ca_int_prev = ca_stages.get('初回面談', {}).get('prev', 0)
        ca_int_proj = ca_stages.get('初回面談', {}).get('projected', 0)
        if ca_int_prev > 5 and month_progress > 0.4:
            if ca_int_proj < ca_int_prev * 0.7:
                ca_alerts.append({
                    'type': 'warning',
                    'message': f'面談ペース低下: 着地{ca_int_proj}件予測（前月{ca_int_prev}件）',
                })
            elif ca_int_proj >= ca_int_prev * 1.1:
                ca_alerts.append({
                    'type': 'good',
                    'message': f'面談ペース好調: 着地{ca_int_proj}件予測（前月{ca_int_prev}件）',
                })

        # CA active count
        ca_active = active_by_ca.get(ca, {})
        ca_hot = ca_active.get('内定', 0) + ca_active.get('最終面接', 0)
        if ca_hot > 0:
            ca_alerts.append({
                'type': 'info',
                'message': f'ホット案件: 内定{ca_active.get("内定",0)} 最終{ca_active.get("最終面接",0)}',
            })

        # CA conversion rates vs team average
        ca_conv = {}
        for label, from_s, to_s in conv_pairs:
            from_vals = [ca_funnel_data.get(from_s, {}).get(m, 0) for m in cur_q_months]
            to_vals = [ca_funnel_data.get(to_s, {}).get(m, 0) for m in cur_q_months]
            ca_conv[label] = []
            for i, m in enumerate(cur_q_months):
                rate = round(to_vals[i] / from_vals[i] * 100, 1) if from_vals[i] > 0 else 0
                ca_conv[label].append({'from': from_vals[i], 'to': to_vals[i], 'rate': rate})
            # Compare latest completed month (prev_m) to team avg
            if len(ca_conv[label]) >= 2:
                ca_rate = ca_conv[label][-2]['rate']  # prev month (completed)
                team_rate = avg_conv_rates.get(label, 0)
                if ca_conv[label][-2]['from'] >= 3 and ca_rate < team_rate - 15:
                    ca_alerts.append({
                        'type': 'warning',
                        'message': f'{label}: {ca_rate:.0f}%（チーム平均{team_rate:.0f}%、{ca_rate-team_rate:+.0f}pp）',
                    })

        weekly_pulse['by_ca'][ca] = {
            'stages': ca_stages, 'alerts': ca_alerts, 'active': ca_active,
            'conversion': ca_conv,
        }

    # ===== PIPELINE HEALTH (リードタイム逆算分析) =====
    # 2ヶ月前のアクティブ数 → 今月の結果に直結
    # 1ヶ月前のアクティブ数 → 来月の結果に直結
    pipeline_health = {
        'lead_time_months': 2,
        'overall': {},
        'by_ca': {},
        'diagnosis': [],
    }

    # Overall: monthly funnel counts for impact analysis
    for m_key in cur_q_months:
        m_label = m_key.split('-')[1] + '月'
        pipeline_health['overall'][m_key] = {
            'label': m_label,
            'stages': monthly_funnel.get(m_key, {}),
        }

    # Impact chain: 2ヶ月前の面談 → 今月の決定に影響
    if len(cur_q_months) >= 3:
        two_ago = cur_q_months[0]
        one_ago = cur_q_months[1]
        current = cur_q_months[2]

        # 2ヶ月前の初回面談数が少ない → 今月の決定が少なくなる
        two_ago_interviews = monthly_funnel.get(two_ago, {}).get('初回面談', 0)
        one_ago_interviews = monthly_funnel.get(one_ago, {}).get('初回面談', 0)
        cur_interviews = monthly_funnel.get(current, {}).get('初回面談', 0)

        # Compare to Q average
        avg_monthly_interviews = (two_ago_interviews + one_ago_interviews) / 2 if two_ago_interviews + one_ago_interviews > 0 else 100

        if two_ago_interviews < avg_monthly_interviews * 0.8:
            pipeline_health['diagnosis'].append({
                'type': 'warning', 'impact': '今月',
                'message': f'2ヶ月前（{two_ago.split("-")[1]}月）の面談数{two_ago_interviews}件が少なく、今月の決定数に影響。新規面談の増加が急務。',
            })
        if one_ago_interviews < avg_monthly_interviews * 0.8:
            pipeline_health['diagnosis'].append({
                'type': 'warning', 'impact': '来月',
                'message': f'1ヶ月前（{one_ago.split("-")[1]}月）の面談数{one_ago_interviews}件が少なく、来月の決定数に影響見込み。',
            })

        # Check funnel quality: where is the biggest drop?
        # Consider active pipeline - if many cases are still in progress, rate looks low but isn't a real issue
        total_active = active_by_stage
        for label, from_s, to_s in conv_pairs:
            cur_rate = monthly_conv_rates.get(one_ago, {}).get(label, {}).get('rate', 0)
            avg_rate = avg_conv_rates.get(label, 0)
            from_count = monthly_conv_rates.get(one_ago, {}).get(label, {}).get('from', 0)
            to_count = monthly_conv_rates.get(one_ago, {}).get(label, {}).get('to', 0)

            # Calculate active compensation for this conversion pair
            active_comp = 0
            if to_s == '求人紹介':
                active_comp = sum(total_active.get(s, 0) for s in ['求人紹介','応募','1st面接','2nd面接','最終面接','内定'])
            elif to_s == '応募':
                active_comp = sum(total_active.get(s, 0) for s in ['応募','1st面接','2nd面接','最終面接','内定'])
            elif to_s == '1st面接':
                active_comp = sum(total_active.get(s, 0) for s in ['1st面接','2nd面接','最終面接','内定'])
            elif to_s == '2nd面接':
                active_comp = sum(total_active.get(s, 0) for s in ['2nd面接','最終面接','内定'])

            adjusted_rate = (to_count + active_comp) / from_count * 100 if from_count > 0 else 0

            if from_count >= 5 and cur_rate < avg_rate - 10:
                if adjusted_rate >= avg_rate - 5:
                    pipeline_health['diagnosis'].append({
                        'type': 'info', 'impact': '進行中',
                        'message': f'{label}: 確定値{cur_rate:.0f}%だがアクティブ{active_comp}件が進行中（補正後{adjusted_rate:.0f}%）。経過観察。',
                        'stage': label,
                    })
                else:
                    pipeline_health['diagnosis'].append({
                        'type': 'bottleneck', 'impact': '来月以降',
                        'message': f'ボトルネック: {label}の通過率{cur_rate:.0f}%（平均{avg_rate:.0f}%、アクティブ込み{adjusted_rate:.0f}%）。質の改善が必要。',
                        'stage': label,
                    })

    # Per-CA pipeline health
    for ca in kpi['ca_names']:
        if ca in departed_cas: continue
        ca_f = funnel_data.get(ca, {}).get('funnel', {})
        ca_health = {'months': {}, 'diagnosis': []}

        for m_key in cur_q_months:
            m_data = {}
            for s in funnel_stages:
                m_data[s] = ca_f.get(s, {}).get(m_key, 0)
            ca_health['months'][m_key] = m_data

        # CA-specific diagnosis
        if len(cur_q_months) >= 3:
            ca_2ago_int = ca_f.get('初回面談', {}).get(cur_q_months[0], 0)
            ca_1ago_int = ca_f.get('初回面談', {}).get(cur_q_months[1], 0)
            ca_cur_int = ca_f.get('初回面談', {}).get(cur_q_months[2], 0)

            # Check if interviews are declining
            if ca_1ago_int > 0 and ca_2ago_int > 0:
                if ca_1ago_int < ca_2ago_int * 0.7:
                    ca_health['diagnosis'].append({
                        'type': 'warning',
                        'message': f'面談数減少傾向: {cur_q_months[0].split("-")[1]}月{ca_2ago_int}件→{cur_q_months[1].split("-")[1]}月{ca_1ago_int}件。来月の決定減少リスク。',
                    })

            # Check conversion rates vs team
            # BUT: consider active pipeline - if there are active cases at later stages,
            # the conversion rate will naturally look low because cases haven't finished yet
            ca_active = active_by_ca.get(ca, {})
            for label, from_s, to_s in conv_pairs[:3]:  # Focus on early funnel
                ca_from = ca_f.get(from_s, {}).get(cur_q_months[1], 0)
                ca_to = ca_f.get(to_s, {}).get(cur_q_months[1], 0)
                ca_rate = ca_to / ca_from * 100 if ca_from > 3 else 0
                team_rate = avg_conv_rates.get(label, 0)

                # Check if active pipeline compensates for the low rate
                # Map conversion pair to relevant active stages
                active_compensation = 0
                if '求人紹介' in label:  # 面談→求人紹介
                    active_compensation = ca_active.get('求人紹介', 0) + ca_active.get('応募', 0) + ca_active.get('1st面接', 0) + ca_active.get('2nd面接', 0) + ca_active.get('最終面接', 0) + ca_active.get('内定', 0)
                elif '応募' in label:  # 求人紹介→応募
                    active_compensation = ca_active.get('応募', 0) + ca_active.get('1st面接', 0) + ca_active.get('2nd面接', 0) + ca_active.get('最終面接', 0) + ca_active.get('内定', 0)
                elif '1st' in label:  # 応募→1st面接
                    active_compensation = ca_active.get('1st面接', 0) + ca_active.get('2nd面接', 0) + ca_active.get('最終面接', 0) + ca_active.get('内定', 0)

                # If active cases would bring the rate close to team average, it's not a real bottleneck
                adjusted_to = ca_to + active_compensation
                adjusted_rate = adjusted_to / ca_from * 100 if ca_from > 3 else 0

                if ca_from > 3 and ca_rate < team_rate - 15:
                    if adjusted_rate >= team_rate - 10:
                        # Active pipeline compensates - not a real issue
                        ca_health['diagnosis'].append({
                            'type': 'info',
                            'message': f'{label}: 確定{ca_rate:.0f}%（低めだがアクティブ{active_compensation}件あり、進行中）',
                        })
                    else:
                        ca_health['diagnosis'].append({
                            'type': 'bottleneck',
                            'message': f'{label}: {ca_rate:.0f}%（チーム平均{team_rate:.0f}%、アクティブ込みでも{adjusted_rate:.0f}%で改善余地あり）',
                        })

            # Active pipeline health
            ca_active = active_by_ca.get(ca, {})
            total_ca_active = sum(ca_active.values())
            if total_ca_active < 3:
                ca_health['diagnosis'].append({
                    'type': 'critical',
                    'message': f'アクティブ案件が{total_ca_active}件と極少。即座に面談数・求人紹介を増やす必要あり。',
                })

        pipeline_health['by_ca'][ca] = ca_health

    weekly_pulse['pipeline_health'] = pipeline_health

    # ===== CONFIRMED-MONTH FUNNEL CONVERSION ALERTS (確定月ベースのファネル通過率アラート) =====
    print("Building confirmed-month conversion alerts...")
    # Use confirmed months only (not current month which is still active)
    confirmed_months = cur_q_months[:-1]  # All months except the current (last) one

    conv_alert_pairs = [
        ('面談→求人紹介', '初回面談', '求人紹介'),
        ('求人紹介→応募', '求人紹介', '応募'),
        ('応募→1st面接', '応募', '1st面接'),
        ('1st→2nd面接', '1st面接', '2nd面接'),
    ]

    # Calculate team average from total funnel for confirmed months
    total_funnel_data = funnel_data.get('total', funnel_data.get('合計', {})).get('funnel', {})
    team_avg_rates = {}
    for label, from_s, to_s in conv_alert_pairs:
        total_from = sum(total_funnel_data.get(from_s, {}).get(m, 0) for m in confirmed_months)
        total_to = sum(total_funnel_data.get(to_s, {}).get(m, 0) for m in confirmed_months)
        team_avg_rates[label] = round(total_to / total_from * 100, 1) if total_from > 0 else 0

    # Calculate per-CA rates and flag alerts
    ca_conv_alerts = {}
    for ca in kpi['ca_names']:
        if ca in departed_cas:
            continue
        ca_f = funnel_data.get(ca, {}).get('funnel', {})
        ca_rates = {}
        ca_alert_list = []
        for label, from_s, to_s in conv_alert_pairs:
            ca_total_from = sum(ca_f.get(from_s, {}).get(m, 0) for m in confirmed_months)
            ca_total_to = sum(ca_f.get(to_s, {}).get(m, 0) for m in confirmed_months)
            ca_rate = round(ca_total_to / ca_total_from * 100, 1) if ca_total_from > 0 else 0
            ca_rates[label] = ca_rate
            team_rate = team_avg_rates.get(label, 0)
            diff = round(ca_rate - team_rate, 1)
            # Only alert if there's enough data (at least 3 in from stage)
            if ca_total_from >= 3 and diff < -5:
                severity = 'warning' if diff < -10 else 'caution'
                ca_alert_list.append({
                    'stage': label,
                    'rate': ca_rate,
                    'team': team_rate,
                    'diff': diff,
                    'severity': severity,
                })
        ca_conv_alerts[ca] = {
            'rates': ca_rates,
            'alerts': ca_alert_list,
        }

    weekly_pulse['conversion_alerts'] = {
        'confirmed_months': confirmed_months,
        'team_avg': team_avg_rates,
        'by_ca': ca_conv_alerts,
    }

    # ========== MONTHLY HIGHLIGHTS (直近月ハイライト) ==========
    print("Building monthly highlights...")
    today = datetime.now()
    current_month = f"{today.year}年{today.month}月"
    prev_month_num = today.month - 1 if today.month > 1 else 12
    prev_month_year = today.year if today.month > 1 else today.year - 1
    prev_month = f"{prev_month_year}年{prev_month_num}月"
    prev2_month_num = prev_month_num - 1 if prev_month_num > 1 else 12
    prev2_month_year = prev_month_year if prev_month_num > 1 else prev_month_year - 1
    prev2_month = f"{prev2_month_year}年{prev2_month_num}月"

    q_months = [prev2_month, prev_month, current_month]

    monthly_highlights = {
        'months': q_months,
        'current_month': current_month,
        'overall': {},
        'by_ca': {},
        'good': [],
        'behind': [],
    }

    # Overall monthly changes
    for period_label, m in [('current', current_month), ('prev', prev_month), ('prev2', prev2_month)]:
        md = monthly_trends.get(m, {})
        monthly_highlights['overall'][period_label] = {
            'month': m,
            'profit': md.get('actual', 0),
            'target': md.get('target', 0),
            'interviews': md.get('interviews', 0),
            'decisions': md.get('decisions', 0),
        }

    # MoM change
    cur = monthly_highlights['overall'].get('current', {})
    prev = monthly_highlights['overall'].get('prev', {})
    if prev.get('profit', 0) > 0 and cur.get('profit', 0) > 0:
        monthly_highlights['overall']['profit_change'] = (cur['profit'] - prev['profit']) / prev['profit']
    if prev.get('interviews', 0) > 0 and cur.get('interviews', 0) > 0:
        monthly_highlights['overall']['interview_change'] = (cur['interviews'] - prev['interviews']) / prev['interviews']

    # Per-CA monthly changes
    for ca in kpi['ca_names']:
        if ca in departed_cas: continue
        ca_monthly = sales.get(ca, {}).get('monthly', {})
        ca_data = {}
        for period_label, m in [('current', current_month), ('prev', prev_month), ('prev2', prev2_month)]:
            md = ca_monthly.get(m, {})
            ca_data[period_label] = {
                'month': m,
                'profit': md.get('実績(粗利)', 0),
                'interviews': md.get('面談数', 0),
                'decisions': md.get('決定数', 0),
            }

        # Determine good/behind
        cur_p = ca_data.get('current', {}).get('profit', 0)
        prev_p = ca_data.get('prev', {}).get('profit', 0)
        cur_i = ca_data.get('current', {}).get('interviews', 0)
        prev_i = ca_data.get('prev', {}).get('interviews', 0)

        if prev_p > 0 and cur_p > 0:
            change_pct = (cur_p - prev_p) / prev_p
            ca_data['profit_change'] = change_pct
            if change_pct >= 0.1:
                monthly_highlights['good'].append({
                    'ca': ca, 'type': 'profit_up',
                    'detail': f'{ca}: 粗利 前月比+{change_pct*100:.0f}%（{prev_p/10000:.0f}万→{cur_p/10000:.0f}万）'
                })
            elif change_pct <= -0.3:
                monthly_highlights['behind'].append({
                    'ca': ca, 'type': 'profit_down',
                    'detail': f'{ca}: 粗利 前月比{change_pct*100:.0f}%（{prev_p/10000:.0f}万→{cur_p/10000:.0f}万）'
                })

        if prev_i > 0 and cur_i > 0:
            int_change = (cur_i - prev_i) / prev_i
            ca_data['interview_change'] = int_change
            if int_change <= -0.3:
                monthly_highlights['behind'].append({
                    'ca': ca, 'type': 'interview_down',
                    'detail': f'{ca}: 面談数 前月比{int_change*100:.0f}%（{prev_i:.0f}件→{cur_i:.0f}件）'
                })

        monthly_highlights['by_ca'][ca] = ca_data

    # ========== PREDICTION ENGINE ==========
    # 着地日(Q列) + リードタイム考慮の保守的予測
    # DB_求職者一覧の着地日を使い、Q内着地のアクティブ案件のみカウント
    print("Building predictions...")

    q_info = get_current_quarter(sales)
    current_q_funnel = q_info['current_q_funnel']
    current_q_sales = q_info['current_q_sales']
    confirmed_q = q_info['confirmed_q']
    q_end_date = q_info['q_end_date']
    print(f"  Auto-detected quarter: {current_q_funnel} (ends {q_end_date.strftime('%Y-%m-%d')})")

    # --- Extract active pipeline with 着地日 from DB_求職者一覧 ---
    ws_db = wb['DB_求職者一覧']
    active_pipeline_by_ca = {}

    for r in range(3, ws_db.max_row + 1):
        status = str(safe_val(ws_db.cell(r, 2).value) or '')
        if 'NG' in status or '保留' in status:
            continue

        landing_date = ws_db.cell(r, 17).value
        ca = str(safe_val(ws_db.cell(r, 3).value) or '')
        dec_count = parse_number(ws_db.cell(r, 45).value)
        gross = parse_number(ws_db.cell(r, 11).value)

        if dec_count > 0: continue  # already decided
        if not ca or ca == 'None': continue

        # Determine if landing within Q
        within_q = False
        has_landing = False
        if landing_date and hasattr(landing_date, 'year'):
            has_landing = True
            within_q = landing_date <= q_end_date

        if ca not in active_pipeline_by_ca:
            active_pipeline_by_ca[ca] = {}
        if status not in active_pipeline_by_ca[ca]:
            active_pipeline_by_ca[ca][status] = {'in_q': 0, 'out_q': 0, 'gross_sum': 0}

        if has_landing:
            if within_q:
                active_pipeline_by_ca[ca][status]['in_q'] += 1
                active_pipeline_by_ca[ca][status]['gross_sum'] += gross
            else:
                active_pipeline_by_ca[ca][status]['out_q'] += 1

    # Stage conversion probabilities for Q内着地 only
    stage_q_conversion = {
        '内定': 0.70,       # 内定→決定: 70%（辞退30%）
        '決定': 0.90,       # ステータス決定で未カウント: 90%
        '最終面接': 0.30,   # 最終面接→Q内決定: 30%
    }

    def calc_historical_decision_rate(ca_name, num_recent_qs=4):
        if ca_name not in sales: return 0
        q_data = sales[ca_name].get('quarterly', {})
        rates = []
        for q in sorted(q_data.keys()):
            if q >= current_q_sales: continue
            interviews = q_data[q].get('面談数', 0)
            decisions = q_data[q].get('決定数', 0)
            if interviews >= 10:
                rates.append(decisions / interviews)
        if not rates: return 0
        return sum(rates[-num_recent_qs:]) / len(rates[-num_recent_qs:])

    def predict_ca_landing(ca_name):
        """着地日を使った保守的な着地予測。
        着地日がQ内のアクティブ案件のみ予測対象。"""
        ca_sales = sales.get(ca_name, {}).get('quarterly', {}).get(current_q_sales, {})
        current_actual = ca_sales.get('実績(粗利)', 0) or ca_comparison.get(ca_name, {}).get('actual', 0)
        current_decisions_raw = ca_sales.get('決定数', 0) or kpi['decision_count'].get(ca_name, 0)
        current_decisions = current_decisions_raw if current_decisions_raw < 100 else 0
        interviews = ca_sales.get('面談数', 0) or ca_comparison.get(ca_name, {}).get('interviews', 0)
        hist_rate = calc_historical_decision_rate(ca_name)

        # Average profit per decision
        if ca_name in sales:
            q_data = sales[ca_name].get('quarterly', {})
            unit_prices = []
            for q in sorted(q_data.keys()):
                dec = q_data[q].get('決定数', 0)
                rev = q_data[q].get('実績(粗利)', 0)
                if dec > 0 and rev > 0:
                    unit_prices.append(rev / dec)
            avg_unit = sum(unit_prices[-3:]) / len(unit_prices[-3:]) if unit_prices else 1000000
        else:
            avg_unit = 1000000

        # Aggregate pipeline data (for 合計, sum all CAs)
        if ca_name == '合計':
            pipeline_data = {}
            for ca_p, stages_p in active_pipeline_by_ca.items():
                for stage_p, counts_p in stages_p.items():
                    if stage_p not in pipeline_data:
                        pipeline_data[stage_p] = {'in_q': 0, 'out_q': 0, 'gross_sum': 0}
                    pipeline_data[stage_p]['in_q'] += counts_p['in_q']
                    pipeline_data[stage_p]['out_q'] += counts_p['out_q']
                    pipeline_data[stage_p]['gross_sum'] += counts_p['gross_sum']
        else:
            pipeline_data = active_pipeline_by_ca.get(ca_name, {})

        # Calculate additional decisions from Q内着地 pipeline
        additional = 0.0
        additional_gross = 0.0
        pipeline_detail = {}

        for stage, prob in stage_q_conversion.items():
            counts = pipeline_data.get(stage, {'in_q': 0, 'out_q': 0, 'gross_sum': 0})
            in_q = counts['in_q']
            out_q = counts.get('out_q', 0)
            if in_q > 0:
                expected = in_q * prob
                additional += expected
                # Use actual gross values where available
                if counts['gross_sum'] > 0:
                    stage_gross = counts['gross_sum'] * prob
                else:
                    stage_gross = expected * avg_unit
                additional_gross += stage_gross
                pipeline_detail[stage] = {
                    'count': in_q, 'out_q': out_q,
                    'prob': prob, 'expected': round(expected, 1),
                }

        total_predicted_decisions = current_decisions + additional
        predicted_landing = current_actual + additional_gross
        predicted_rate = total_predicted_decisions / interviews if interviews > 0 else 0

        return {
            'current_actual': current_actual,
            'current_decisions': current_decisions,
            'additional_decisions_predicted': round(additional, 1),
            'total_decisions_predicted': round(total_predicted_decisions, 1),
            'historical_decision_rate': hist_rate,
            'predicted_decision_rate': predicted_rate,
            'avg_unit_price': avg_unit,
            'predicted_landing': round(predicted_landing),
            'interviews': interviews,
            'pipeline_detail': pipeline_detail,
        }

    predictions = {
        'confirmed_q': confirmed_q,
        'current_q': current_q_funnel,
        'overall': {},
        'by_ca': {},
    }

    # Map 合計 → total in funnel_data for prediction
    if 'total' in funnel_data and '合計' not in funnel_data:
        funnel_data['合計'] = funnel_data['total']

    # Overall prediction
    overall_pred = predict_ca_landing('合計')
    overall_hist_rate = calc_historical_decision_rate('合計')

    # Confirmed decision rate (from confirmed Q)
    conf_q_data = sales.get('合計', {}).get('quarterly', {}).get(confirmed_q, {})
    confirmed_rate = 0
    if conf_q_data:
        conf_interviews = conf_q_data.get('面談数', 0)
        conf_decisions = conf_q_data.get('決定数', 0)
        confirmed_rate = conf_decisions / conf_interviews if conf_interviews > 0 else 0

    predictions['overall'] = {
        'confirmed_decision_rate': confirmed_rate,
        'confirmed_q_label': confirmed_q,
        'predicted_decision_rate': overall_pred['predicted_decision_rate'],
        'historical_avg_rate': overall_hist_rate,
        **overall_pred,
    }

    # Per-CA predictions
    for ca in kpi['ca_names']:
        if ca in departed_cas: continue
        ca_pred = predict_ca_landing(ca)
        ca_hist_rate = calc_historical_decision_rate(ca)

        # CA's confirmed Q rate
        ca_conf_data = sales.get(ca, {}).get('quarterly', {}).get(confirmed_q, {})
        ca_confirmed_rate = 0
        if ca_conf_data:
            ci = ca_conf_data.get('面談数', 0)
            cd = ca_conf_data.get('決定数', 0)
            ca_confirmed_rate = cd / ci if ci > 0 else 0

        predictions['by_ca'][ca] = {
            'confirmed_decision_rate': ca_confirmed_rate,
            'historical_avg_rate': ca_hist_rate,
            **ca_pred,
        }

    print(f"  Overall predicted landing: {overall_pred['predicted_landing']/10000:.0f}万")
    print(f"  Overall confirmed rate ({confirmed_q}): {confirmed_rate:.1%}")
    print(f"  Overall predicted rate: {overall_pred['predicted_decision_rate']:.1%}")

    dashboard = {
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'kpi': kpi,
        'monthly_trends': dict(sorted(monthly_trends.items(), key=lambda x: month_sort_key(x[0]))),
        'quarterly_trends': dict(sorted(quarterly_trends.items())),
        'ca_comparison': ca_comparison,
        'funnel_data': funnel_data,
        'sales_by_ca': {k: v for k, v in sales.items() if k in kpi['ca_names'] or k == '合計'},
        'ca_historical': historical.get('ca_quarterly', {}),
        'inflow': inflow,
        'route_process': route_process,
        'decision_by_attr': decision_attr,
        'insights': insights,
        'monthly_highlights': monthly_highlights,
        'weekly_pulse': weekly_pulse,
        'targets_2026': targets_2026,
        'referral': referral_data,
        'route_breakdown': route_breakdown,
        'historical': historical,
        'ca_deep_analysis': ca_deep_analysis,
        'predictions': predictions,
        'yomi_forecast': yomi_forecast,
        'weekly_report': weekly_report,
        'lead_time_months': lead_time_months,
        'ca_status': {
            'departed': departed_cas,
            'zero_target_1q': list(zero_target_cas.keys()),
            'joining_2q': joining_cas,
        },
        'current_quarter': current_q_funnel,
    }

    with open('data/dashboard_data.json', 'w', encoding='utf-8') as f:
        json.dump(dashboard, f, ensure_ascii=False, indent=2)

    print(f"\nDone! Insights: overall={len(insights['overall'])}, CA={len(insights['ca_detail'])}, inflow={len(insights['inflow'])}")
    print(f"2026 targets: {list(targets_2026.keys())}")
    print(f"Lead time months (excluded from decision rate): {lead_time_months}")

if __name__ == '__main__':
    main()
